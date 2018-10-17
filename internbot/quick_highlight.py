from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# need to write here
highlight_style = PatternFill("solid", fgColor="C00201")
font_style = Font(name='Arial', size=10, color='ffffff')

workbook = load_workbook("/Users/y2analytics/Downloads/Unhighlighted.xlsx")

for sheet in workbook.worksheets:
    if sheet.title != 'TOC':
        label_column = sheet['A']
        response_column = sheet['B']
        for cell in label_column:
            if cell.row > 3 and \
                cell.value is not None and \
                cell.value != "Results are based on two-sided tests. For each significant pair, the key of the category with the smaller column proportion appears in the category with the larger column proportion.\n Significance level for upper case letters (A, B, C): .05" and \
                cell.value != "a. This category is not used in comparisons because its column proportion is equal to zero or one." and \
                cell.value != "b. This category is not used in comparisons because the sum of case weights is less than two." and \
                cell.value != "Comparisons of Column Proportions":
                current_label = cell.value
                for response_cell in response_column:
                    if response_cell.value is not None and \
                        response_cell.value != "Signma" and \
                        response_cell.value != "Total":
                        response_label = current_label+response_cell.value
                        print response_label