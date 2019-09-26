from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
import os

class Spreadsheet(object):

    def __init__(self, is_qualtrics, resources_filepath):
        self.__workbook = Workbook()
        self.__is_qualtrics = is_qualtrics
        self.resources_filepath = resources_filepath

        if is_qualtrics:
            # fill colors
            self.__header_fill = PatternFill("solid", fgColor = "1E262E")
        else:
            self.__header_fill = PatternFill("solid", fgColor = "0F243E")

        self.__table_fill = PatternFill("solid", fgColor = "E7E6E6")
       
        # font styles
        self.__font_reg = Font(name = 'Arial', size = 8)
        self.__font_bold = Font(name = 'Arial', size = 8, bold = True)
        self.__font_back_hyperlink = Font(name = 'Arial', size = 8, color = "A2AAAD")

        # alignments
        self.__align_center = Alignment(horizontal = "center", vertical = "center", wrapText = True)
        self.__align_left = Alignment(horizontal = "left", vertical = "center", wrapText = True)

        # borders
        self.__thin_bottom = Border(bottom = Side(style = 'thin'))

    def write_appendix(self, questions):
        if self.__workbook.get_sheet_by_name("Sheet") is not None:
            toc_sheet = self.__workbook.get_sheet_by_name('Sheet')
            toc_sheet.title = "TOC"
        else:
            toc_sheet = self.__workbook.create_sheet("TOC")

        self.write_toc(toc_sheet, questions)

        for question, value in questions.items():
            new_sheet = self.__workbook.create_sheet(question)
            self.write_question(value, new_sheet)

    def write_toc(self, sheet, questions):
        print("Writing TOC")
        self.write_header(sheet)

        sheet["A2"].value = "Question Name"
        sheet["A2"].font = self.__font_bold
        sheet["A2"].alignment = self.__align_center

        sheet["B2"].value = "Question Title"
        sheet["B2"].font = self.__font_bold
        sheet["B2"].alignment = self.__align_center

        sheet["C2"].value = "Verbatim response count"
        sheet["C2"].font = self.__font_bold
        sheet["C2"].alignment = self.__align_center

        current_row = 3
        for question, value in questions.items():
            question_name = "A%s" % str(current_row)
            question_title = "B%s" % str(current_row)
            base_size = "C%s" % str(current_row)

            sheet[question_name].value = "=HYPERLINK(\"#'%s'!A1\",\"%s\")" % (question, question)
            sheet[question_name].font = Font(name = 'Arial', size = 8, underline = 'single')
            sheet[question_name].alignment = self.__align_center
            sheet[question_name].border = self.__thin_bottom

            sheet[question_title].value = value.prompt
            sheet[question_title].font = self.__font_reg
            sheet[question_title].alignment = self.__align_center
            sheet[question_title].border = self.__thin_bottom

            sheet[base_size].value = value.response_count
            sheet[base_size].font = self.__font_reg
            sheet[base_size].alignment = self.__align_center
            sheet[base_size].border = self.__thin_bottom
            current_row += 1

    def write_header(self, sheet):
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 55
        sheet.column_dimensions["C"].width = 33

        if self.__is_qualtrics:
            sheet.row_dimensions[1].height = 35
        else:
            sheet.row_dimensions[1].height = 52

        sheet["A1"].fill = self.__header_fill
        sheet["B1"].fill = self.__header_fill
        sheet["C1"].fill = self.__header_fill

        if self.__is_qualtrics:
            logo = Image(os.path.join(self.resources_filepath, "QLogo.png"))
        else:
            logo = Image(os.path.join(self.resources_filepath, "y2_xtabs.png"))
        sheet.add_image(logo, "C1")

    def write_question(self, question, sheet):
        to_print = "Writing: %s" % question.name
        print(to_print)
        self.write_header(sheet)
        sheet.row_dimensions[2].height = 36

        sheet["A1"].value = '=HYPERLINK("#TOC!A1","Return to Table of Contents")'
        sheet["A1"].font = self.__font_back_hyperlink
        sheet["A1"].alignment = self.__align_center

        sheet.merge_cells(start_column=1, end_column=3, start_row=2, end_row=2)

        sheet["A2"].fill = self.__table_fill
        sheet["B2"].fill = self.__table_fill
        sheet["C2"].fill = self.__table_fill

        sheet["A2"].value = "%s: %s" % (question.name, question.prompt)
        sheet["A2"].alignment = self.__align_center
        sheet["A2"].font = self.__font_reg

        self.write_responses(question.responses, sheet)

    def write_responses(self, responses, sheet):
        current_row = 3
        for response in responses:
            sheet.row_dimensions[current_row].height = 25
            a_cell = "A%s" % str(current_row)
            b_cell = "B%s" % str(current_row)
            c_cell = "C%s" % str(current_row)
            sheet.merge_cells(start_row=current_row, end_row=current_row, start_column = 1, end_column = 3)

            sheet[a_cell].value = response
            sheet[a_cell].font = self.__font_reg
            sheet[a_cell].alignment = self.__align_left
            sheet[a_cell].border = self.__thin_bottom
            sheet[b_cell].border = self.__thin_bottom
            sheet[c_cell].border = self.__thin_bottom

            current_row += 1

    def save(self, path_to_output):
        self.__workbook.save(path_to_output)
