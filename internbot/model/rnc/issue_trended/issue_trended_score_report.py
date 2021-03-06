from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from collections import OrderedDict
   
class IssueTrendedReport(object):

    def __init__ (self, models, number_of_rounds):
        # create final workbook
        self.models = models
        self.workbook = Workbook()
        self.total_count = 0
        self.rounds = int(number_of_rounds)

        # font details
        self.titles_style = Font(name = 'Calibri (Body)', size = 11, bold = True)
        self.general_style = Font(name = 'Calibri (Body)', size = 11)
        self.second_row = Font(name = 'Calibri (Body)', size = 11, bold = True, italic = True)

        # border details
        self.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        self.top_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'))
        self.middle_border = Border(left=Side(style='thick'), right=Side(style='thick'))
        self.bottom_border = Border(left=Side(style='thick'), right=Side(style='thick'), bottom=Side(style='thick'))

        # fill/highlight details
        self.grey = PatternFill("solid", fgColor="B7B7B7")
        self.key_variables = PatternFill("solid", fgColor="FFC5C5")
        self.key_models = PatternFill("solid", fgColor="D4D4D4")

        # calculate the excel alphabet from G to ZZ
        alphabet = []
        for letter in range(71, 91):
            alphabet.append(chr(letter))

        self.extend_alphabet = []
        self.extend_alphabet.extend(alphabet)
        complete_alphabet = []
        for letter in range(65, 91):
            complete_alphabet.append(chr(letter))
        index = 0
        while index < len(complete_alphabet):
            for letter in complete_alphabet:
                double_letters = "%s%s" % (complete_alphabet[index], letter)
                self.extend_alphabet.append(double_letters)
            index += 1

        # write report
        self.build_sheets()

    def build_sheets(self):
        model_names = self.models.list_model_names()
        default_sheet = False
        if self.workbook.get_sheet_by_name("Sheet") is not None:
            default_sheet = True

        alternate_tabs = True
        for model_name in model_names:
            title = model_name[:30]
            if default_sheet is True:
                sheet = self.workbook.get_sheet_by_name("Sheet")
                sheet.title = title
                default_sheet = False
            else:
                sheet = self.workbook.create_sheet(title)

            current = self.models.get_model(model_name)
            self.write_first_row(sheet)
            self.write_fields(sheet, current)
            self.write_differences(sheet, current)

            if alternate_tabs is True:
                sheet.sheet_properties.tabColor = '7F7F7F'
                alternate_tabs = False
            else:
                sheet.sheet_properties.tabColor = 'A6494A'
                alternate_tabs = True

    def write_first_row(self, current_sheet):
        current_sheet.row_dimensions[1].height = 30

        current_sheet["A1"].value = "Field Name"
        current_sheet["A1"].font = self.titles_style
        current_sheet["A1"].border = self.border
        current_sheet["A1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["A"].width = 33

        current_sheet["B1"].value = "Grouping"
        current_sheet["B1"].font = self.titles_style
        current_sheet["B1"].border = self.border
        current_sheet["B1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["B"].width = 29

        current_sheet["C1"].value = "Count"
        current_sheet["C1"].font = self.titles_style
        current_sheet["C1"].border = self.border
        current_sheet["C1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["C"].width = 11

        current_sheet["D1"].value = "Percent"
        current_sheet["D1"].font = self.titles_style
        current_sheet["D1"].border = self.border
        current_sheet["D1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["D"].width = 11

        current_sheet["E1"].value = "Current Round - Previous Round"
        current_sheet["E1"].font = self.titles_style
        current_sheet["E1"].border = self.border
        current_sheet["E1"].alignment = Alignment(horizontal="center", wrap_text = True)

        current_sheet.column_dimensions["E"].width = 13

        current_sheet["F1"].value = "Current Round - First Round"
        current_sheet["F1"].font = self.titles_style
        current_sheet["F1"].border = self.border
        current_sheet["F1"].alignment = Alignment(horizontal="center", wrap_text = True)

        current_sheet.column_dimensions["F"].width = 13

        round_iteration = self.rounds
        index = 0
        while round_iteration > 0:
            header = "Round %s %s" % (round_iteration, self.models.round_date(round_iteration))
            column = self.extend_alphabet[index]
            cell = "%s1" % column
            current_sheet[cell].value = header
            current_sheet[cell].border = self.border
            current_sheet[cell].font = self.titles_style
            current_sheet[cell].alignment = Alignment(horizontal="center", wrap_text = True)
            round_iteration -= 1
            index += 1

    def write_fields(self, current_sheet, model):
        current_row_A = 2
        current_row_B = 2
        field_names = model.list_field_names()
        for field_name in field_names:
            location = "A%s" % current_row_A
            field = model.get_field(field_name)
            current_sheet[location].value = field.name
            current_sheet[location].font = self.titles_style
            current_sheet[location].alignment = Alignment(horizontal="center")

            # merge
            group_count = len(field.list_grouping_names())
            end = (current_row_A + group_count) - 1
            current_sheet.merge_cells(start_column=1, end_column=1, start_row=current_row_A, end_row = end)
            current_row_A += group_count
            
            # write groups
            current_row_B = self.write_groupings(current_sheet, current_row_B, field)

    def write_groupings(self, current_sheet, current_row, field):
        grouping_names = field.list_grouping_names()
        middle_border = False
        top_border = True
        if len(grouping_names) > 2:
            middle_border = True
        for grouping_name in grouping_names:
            field_cell = "A%s" % (current_row)
            grouping_cell = "B%s" % (current_row)
            count_cell = "C%s" % (current_row)
            percent_cell = "D%s" % (current_row)
            prev_diff_cell = "E%s" % (current_row)
            first_diff_cell = "F%s" % (current_row)

            grouping = field.get_grouping(grouping_name)

            if grouping_name == "All Voters":
                self.total_count = grouping.count
                current_sheet[field_cell].fill = self.grey
                current_sheet[field_cell].font = self.second_row
                current_sheet[grouping_cell].fill = self.grey
                current_sheet[grouping_cell].font = self.second_row
                current_sheet[count_cell].fill = self.grey
                current_sheet[count_cell].font = self.second_row
                current_sheet[percent_cell].fill = self.grey
                current_sheet[percent_cell].font = self.second_row
                current_sheet[prev_diff_cell].font = self.second_row
                current_sheet[first_diff_cell].font = self.second_row

            current_sheet[grouping_cell].value = grouping_name
            current_sheet[grouping_cell].font = self.titles_style
            current_sheet[count_cell].value = grouping.count
            current_sheet[percent_cell].value = (grouping.count)/float(self.total_count)
            current_sheet[percent_cell].number_format = '0%'
            current_sheet[prev_diff_cell].number_format = '0%'
            current_sheet[first_diff_cell].number_format = '0%'
            current_sheet[count_cell].number_format = '#,##0'

            index = 0
            round_iteration = self.rounds
            while round_iteration > 0:
                round_cell = "%s%s" % (self.extend_alphabet[index], current_row)
                frequency = grouping.round_frequency(round_iteration)
                if frequency == "NA" or frequency == "":
                    current_sheet[round_cell].fill = self.grey
                else:
                    current_sheet[round_cell].value = float(frequency)
                    current_sheet[round_cell].number_format = '0%'
                if current_row == 2:
                    current_sheet[round_cell].font = self.second_row
                    current_sheet[round_cell].fill = self.grey
                if top_border is True:
                    current_sheet[round_cell].border = self.top_border
                elif middle_border is True:
                    current_sheet[round_cell].border = self.middle_border
                round_iteration -= 1
                index += 1

            if top_border is True:
                current_sheet[field_cell].border = self.top_border
                current_sheet[grouping_cell].border = self.top_border
                current_sheet[count_cell].border = self.top_border
                current_sheet[percent_cell].border = self.top_border
                current_sheet[prev_diff_cell].border = self.top_border
                current_sheet[first_diff_cell].border = self.top_border
                top_border = False
            elif middle_border is True:
                current_sheet[field_cell].border = self.middle_border
                current_sheet[grouping_cell].border = self.middle_border
                current_sheet[count_cell].border = self.middle_border
                current_sheet[percent_cell].border = self.middle_border
                current_sheet[prev_diff_cell].border = self.middle_border
                current_sheet[first_diff_cell].border = self.middle_border

            current_row += 1

        last_row = current_row - 1
        field_cell = "A%s" % (last_row)
        grouping_cell = "B%s" % (last_row)
        count_cell = "C%s" % (last_row)
        percent_cell = "D%s" % (last_row)
        prev_diff_cell = "E%s" % (last_row)
        first_diff_cell = "F%s" % (last_row)
        current_sheet[field_cell].border = self.bottom_border
        current_sheet[grouping_cell].border = self.bottom_border
        current_sheet[count_cell].border = self.bottom_border
        current_sheet[percent_cell].border = self.bottom_border
        current_sheet[prev_diff_cell].border = self.bottom_border
        current_sheet[first_diff_cell].border = self.bottom_border
        
        index = 0
        round_iteration = self.rounds
        while round_iteration > 0:
            round_cell = "%s%s" % (self.extend_alphabet[index], last_row)
            current_sheet[round_cell].border = self.bottom_border
            round_iteration -= 1
            index += 1

        return current_row
 
    def write_differences(self, score_sheet, current):
        current_row = 2
        field_names = current.list_field_names()
        for field_name in field_names:
            field = current.get_field(field_name)
            grouping_names = field.list_grouping_names()
            for grouping_name in grouping_names:
                prev_diff_cell = "E%s" % (current_row)
                first_diff_cell = "F%s" % (current_row)

                if self.rounds == 1:
                    score_sheet[prev_diff_cell].value = "--%"
                    score_sheet[first_diff_cell].value = "--%"
                else:
                    current_round_cell = "G%s" % (current_row)
                    previous_round_cell = "H%s" % (current_row)
                    first_round_cell = self.find_first_round(score_sheet, current, current_row)
                    
                    # check if first round is nonexistent
                    no_first = False
                    if first_round_cell == previous_round_cell:
                        no_first = True

                    # check if previous round is empty
                    no_prev = False
                    if score_sheet[previous_round_cell].value is None:
                        no_prev = True

                    # check if current round is empty
                    no_current = False
                    if score_sheet[current_round_cell].value is None:
                        no_current = True

                    ### figure out edge cases here ###
                    ## a model with no current frequency -- a model taken off current round
                    if no_current == True:
                        score_sheet[prev_diff_cell].value = "--%"
                        score_sheet[first_diff_cell].value = "--%"
                    ## a model only added to the current round ##
                    elif no_prev == True and no_first == True:
                        score_sheet[prev_diff_cell].value = "--%"
                        score_sheet[first_diff_cell].value = "--%"
                    ## a model in first and current round only ##
                    elif no_prev == True and no_first == False:
                        score_sheet[prev_diff_cell].value = "--%"
                        score_sheet[first_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[first_round_cell].value)
                        score_sheet[first_diff_cell].value = "=%s - %s" % (current_round_cell, first_round_cell)
                        score_sheet[first_diff_cell].number_format = '0%'
                    ## a model in previous and current round only
                    elif no_prev == False and no_first == True:
                        score_sheet[first_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[first_round_cell].value)
                        score_sheet[first_diff_cell].value = "=%s - %s" % (current_round_cell, first_round_cell)
                        score_sheet[first_diff_cell].number_format = '0%'
                        score_sheet[prev_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[previous_round_cell].value)
                        score_sheet[prev_diff_cell].value = "=%s - %s" % (current_round_cell, previous_round_cell)
                        score_sheet[prev_diff_cell].number_format = '0%'
                    ## no current edge cases apply ##
                    else:
                        score_sheet[prev_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[previous_round_cell].value)
                        score_sheet[prev_diff_cell].value = "=%s - %s" % (current_round_cell, previous_round_cell)
                        score_sheet[prev_diff_cell].number_format = '0%'
                        score_sheet[first_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[first_round_cell].value)
                        score_sheet[first_diff_cell].value = "=%s - %s" % (current_round_cell, first_round_cell)
                        score_sheet[first_diff_cell].number_format = '0%'
                current_row += 1
            
    def find_first_round(self, score_sheet, model, current_row):
        index = self.rounds - 1
        first_cell = "H%s" % current_row
        while index > 1:
            first_column = self.extend_alphabet[index]
            test_cell = "%s%s" % (first_column, current_row)
            if score_sheet[test_cell].value == None:
                index -= 1
            else:
                return test_cell
        return first_cell 

    def highlight(self, first_value, second_value):
        darkest_negative = PatternFill("solid", fgColor="B80001")
        medium_dark_negative = PatternFill("solid", fgColor="CD4748")
        medium_negative = PatternFill("solid", fgColor="DF8A8C")
        medium_light_negative = PatternFill("solid", fgColor="EAB9BB")
        lightest_negative = PatternFill("solid", fgColor="F1D2D5")
 
        darkest_positive = PatternFill("solid", fgColor="02A747")
        medium_dark_positive = PatternFill("solid", fgColor="2AB666")
        medium_positive = PatternFill("solid", fgColor="5AC88B")
        medium_light_positive = PatternFill("solid", fgColor="B6E7CE")
        lightest_positive = PatternFill("solid", fgColor="E6F6F0")
        
        grey = PatternFill("solid",fgColor="E6E6E6")
        if first_value is None or second_value is None:
            return grey
        
        result = first_value - second_value
        if result < 0:
            if result >= -0.01:
                return lightest_negative
            elif result >= -0.03:
                return medium_light_negative
            elif result >= -0.05:
                return medium_negative
            elif result >= -0.07:
                return medium_dark_negative
            elif result < -0.07:
                return darkest_negative
            else:
                return grey
        elif result > 0:
            if result <= 0.01:
                return lightest_positive
            elif result <= 0.03:
                return medium_light_positive
            elif result <= 0.05:
                return medium_positive
            elif result <= 0.07:
                return medium_dark_positive
            elif result > 0.07:
                return darkest_positive
            else:
                return grey
        else:
            return grey

    def save(self, path_to_output):
        self.workbook.save(path_to_output)
        