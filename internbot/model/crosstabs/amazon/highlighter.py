from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from model.crosstabs.amazon.cell import Cells, Cell, PercentageCell, PopulationCell, SignificantMarker

class Highlighter(object):

    def __init__(self):
        self.groups = {}
        self.responses = {}
        self.__cells = Cells()
        self.highlight_style = PatternFill("solid", fgColor="C00201")
        self.font_style = Font(name='Arial', size=10, color='ffffff')
    
    def highlight(self, workbook, path_to_output, is_trended_amazon = True):
        self.workbook = workbook
        self.trended = is_trended_amazon
        for sheet in self.workbook.worksheets:
            if sheet.title != 'TOC':
                self.parse_rows(sheet)
                self.parse_columns(sheet)
                self.create_cells()
                self.assign_significant(sheet)
                self.highlight_significant(sheet)
            self.groups = {}
            self.responses = {}
            self.__cells = Cells()
        print("Finished!")
        self.workbook.save(path_to_output)
        
    def parse_rows(self, sheet):
        response_column = sheet['B']
        for cell in response_column:
            if cell.value is not None and \
               cell.value != 'Sigma' and \
               cell.value != 'Total' and \
               cell.value not in self.responses.keys():
                self.responses[cell.value] = cell.row
            elif cell.value is not None and \
                cell.value in self.responses.keys():
                temp_storage = self.responses.get(cell.value)
                self.responses[cell.value] = [temp_storage]
                self.responses[cell.value].append(cell.row)

    def parse_columns(self, sheet):
        if self.trended == True:
            group_row = sheet['5']
        else:
            group_row = sheet['3']
        iteration = 1
        for cell in group_row:
            if cell.value is not None:
                if cell.value in self.groups.keys():
                    new_name = "%s_%s" % (cell.value, iteration)
                    iteration += 1
                    self.groups[new_name] = cell.column_letter
                else:
                    self.groups[cell.value] = cell.column_letter

    def create_cells(self):
        for group in self.groups:
            for label in self.responses:
                location1 = str(self.groups[group]) + str(self.responses[label][0])
                location2 = str(self.groups[group]) + str(self.responses[label][0] + 1)
                location3 = str(self.groups[group]) + str(self.responses[label][1])
                
                self.__cells.add(PopulationCell(label, group, location1))
                self.__cells.add(PercentageCell(label, group, location2))
                self.__cells.add(SignificantMarker(label, group, location3))

    def assign_significant(self, sheet):
        for cell in self.__cells:
            if cell.type == 'SignificantMarker':
                sheet_cell = sheet[cell.location].value
                if sheet_cell is not None and sheet_cell.isupper() is True:
                    match_cell = self.__cells.matching_cells(cell)
                    match_cell.is_significant = True

    def highlight_significant(self, sheet):
        for cell in self.__cells:
            if cell.is_significant is True:
                sheet[cell.location].fill = self.highlight_style
                sheet[cell.location].font = self.font_style