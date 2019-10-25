from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import range_boundaries
from collections import OrderedDict
import os
import re

class Formatter(object):

    def __init__(self):
        self.__tables = OrderedDict()
        self.__stop_report = False

        self.set_fonts()
        self.set_alignments()
        self.set_borders()
        self.calculate_excel_col_names()

    def set_fonts(self):
        self.__font_reg = Font(name = 'Arial', size = 8)
        self.__font_small = Font(name = 'Arial', size = 7)
        self.__font_bold = Font(name = 'Arial', size = 8, bold = True)
        self.__font_back_hyperlink = Font(name = 'Arial', size = 8, color = "A2AAAD")
        self.__font_toc_table = Font(name = 'Arial', size = 8, underline = 'single')

    def set_alignments(self):
        self.__align_banners = Alignment(horizontal = "center", vertical = "bottom", wrapText = True)
        self.__align_names = Alignment(horizontal = "left", vertical = "center", wrapText = True)
        self.__align_center = Alignment(horizontal = "center", vertical = "center", wrapText = True)
        self.__align_left = Alignment(horizontal = "left", vertical = "center", wrapText = True)
        self.__align_top = Alignment(horizontal = "left", vertical = "top", wrapText = True)

    def set_borders(self):
        self.__thin_bottom = Border(bottom = Side(style = 'thin'))
        self.__thick_top = Border(top = Side(style = 'thick'))
        self.__thin_top = Border(top = Side(style = 'thin'))
        self.__thick_left = Border(left = Side(style = 'thick'))

    def calculate_excel_col_names(self):
        # calculate the excel alphabet from A to ZZZ
        alphabet = []
        for letter in range(65, 91):
            alphabet.append(chr(letter))
    
        self.__extend_alphabet = []
        self.__extend_alphabet.extend(alphabet)
        full_alphabet = []
        double_alphabet = []
        for letter in range(65, 91):
            full_alphabet.append(chr(letter))
        index = 0
        while index < len(full_alphabet):
            for letter in full_alphabet:
                double_letters = "%s%s" % (full_alphabet[index], letter)
                double_alphabet.append(double_letters)
                self.__extend_alphabet.append(double_letters)
            index += 1
        index = 0
        while index < len(double_alphabet):
            for letter in full_alphabet:
                triple_letters = "%s%s" % (double_alphabet[index], letter)
                self.__extend_alphabet.append(triple_letters)
            index += 1

    def format_qresearch_report(self, path_to_workbook, image_path):
        table_index = 0
        self.__workbook = load_workbook(path_to_workbook)
        print("Loading workbook")

        self.__image_path = image_path
        self.set_template()

        for sheet in self.__workbook.worksheets:
            if sheet.title == 'TOC':
                print("Parsing TOC")
                self.parse_toc(sheet)
            else:
                if self.__stop_report:
                    return
                else:
                    items = list(self.__tables.items())
                    title = items[table_index][0]
                    sheet.title = title
                    to_print = "Formatting: %s" % sheet.title
                    print(to_print)
                    self.format_sheet(sheet)
                    table_index += 1    
        print("Done!")

    def set_template(self):
        if (os.path.basename(self.__image_path) == "QLogo.png"):
            self.__row_height = 35
            self.__header_fill = PatternFill("solid", fgColor = "1E262E")
            self.__hi_significant_fill = PatternFill("solid", fgColor = "2DCCD3")
            self.__lo_significant_fill = PatternFill("solid", fgColor = "2DCCD3")
        elif (os.path.basename(self.__image_path) == "y2_xtabs.png"):
            self.__row_height = 52
            self.__header_fill = PatternFill("solid", fgColor = "0F243E")
            self.__hi_significant_fill = PatternFill("solid", fgColor = "2083E7")
            self.__lo_significant_fill = PatternFill("solid", fgColor = "2083E7")
        elif (os.path.basename(self.__image_path) == "y2_utpol_logo.png"):
            self.__row_height = 52
            self.__header_fill = PatternFill("solid", fgColor = "0F243E")
            self.__hi_significant_fill = PatternFill("solid", fgColor = "2083E7")
            self.__lo_significant_fill = PatternFill("solid", fgColor = "2083E7")
        elif (os.path.basename(self.__image_path) == "whatsapp.png"):
            self.__row_height = 57
            self.__header_fill = PatternFill("solid", fgColor = "445963")
            self.__hi_significant_fill = PatternFill("solid", fgColor = "1EBDA5")
            self.__lo_significant_fill = PatternFill("solid", fgColor = "1EBDA5")

        self.__table_fill = PatternFill("solid", fgColor = "E7E6E6")
        self.__white_fill = PatternFill("solid", fgColor = "FFFFFF")

    def parse_toc(self, toc_sheet):
        table_no_col = "A"
        table_title_col = "B"
        base_desc_col = "C"
        base_size_col = "D"

        table_count = len(self.__workbook.worksheets) - 1

        current_row = 1

        table_no_cell = "%s%s" % (table_no_col, str(current_row))
        table_title_cell = "%s%s" % (table_title_col, str(current_row))
        base_desc_cell = "%s%s" % (base_desc_col, str(current_row))
        base_size_cell = "%s%s" % (base_size_col, str(current_row))

        toc_sheet.column_dimensions[table_no_col].width = 9
        toc_sheet.column_dimensions[table_title_col].width = 100
        toc_sheet.column_dimensions[base_desc_col].width = 33
        toc_sheet.column_dimensions[base_size_col].width = 34

        if toc_sheet["A1"].value is not None:
            toc_sheet.insert_rows(1)

        toc_sheet.row_dimensions[1].height = self.__row_height
        
        toc_sheet[table_no_cell].fill = self.__header_fill
        toc_sheet[table_title_cell].fill = self.__header_fill
        toc_sheet[base_desc_cell].fill = self.__header_fill
        toc_sheet[base_size_cell].fill = self.__header_fill

        logo = Image(self.__image_path)
        toc_sheet.add_image(logo, base_size_cell)
        
        current_row += 1
        table_no_cell = "%s%s" % (table_no_col, str(current_row))
        while toc_sheet[table_no_cell].value is not None:
            table_title_cell = "%s%s" % (table_title_col, str(current_row))
            base_desc_cell = "%s%s" % (base_desc_col, str(current_row))
            base_size_cell = "%s%s" % (base_size_col, str(current_row))

            if current_row == 1:
                toc_sheet[table_no_cell].fill = self.__header_fill
                toc_sheet[table_title_cell].fill = self.__header_fill
                toc_sheet[base_desc_cell].fill = self.__header_fill
                toc_sheet[base_size_cell].fill = self.__header_fill

                logo = Image(self.__image_path)
                toc_sheet.add_image(logo, base_size_cell)

            elif current_row == 2:
                toc_sheet[table_no_cell].font = self.__font_bold
                toc_sheet[table_no_cell].value = "Table #"
                toc_sheet[table_no_cell].alignment = self.__align_center

                toc_sheet[table_title_cell].font = self.__font_bold
                toc_sheet[table_title_cell].value = "Question Title"
                toc_sheet[table_title_cell].alignment = self.__align_center

                toc_sheet[base_desc_cell].font = self.__font_bold
                toc_sheet[base_desc_cell].value = "Base Description"
                toc_sheet[base_desc_cell].alignment = self.__align_center

                toc_sheet[base_size_cell].font = self.__font_bold
                toc_sheet[base_size_cell].value = "Base Size (N count)"
                toc_sheet[base_size_cell].alignment = self.__align_center

            else:
                table = TOCTable(toc_sheet[table_no_cell].value)

                toc_sheet[table_no_cell].font = self.__font_toc_table
                table_name = toc_sheet[table_no_cell].value
                toc_sheet[table_no_cell].value = "=HYPERLINK(\"#'%s'!A1\",\"%s\")" % (table_name, table_name)
                toc_sheet[table_no_cell].alignment = self.__align_center
                toc_sheet[table_no_cell].border = self.__thin_bottom

                table.prompt = toc_sheet[table_title_cell].value
                toc_sheet[table_title_cell].font = self.__font_reg
                toc_sheet[table_title_cell].alignment = self.__align_left
                toc_sheet[table_title_cell].border = self.__thin_bottom

                table.description = toc_sheet[base_desc_cell].value
                toc_sheet[base_desc_cell].font = self.__font_reg
                toc_sheet[base_desc_cell].alignment = self.__align_left
                toc_sheet[base_desc_cell].border = self.__thin_bottom

                table.size = toc_sheet[base_size_cell].value
                toc_sheet[base_size_cell].font = self.__font_reg
                toc_sheet[base_size_cell].alignment = self.__align_left
                toc_sheet[base_size_cell].border = self.__thin_bottom

                self.__tables[table.name] = table
            
            current_row += 1
            table_no_cell = "%s%s" % (table_no_col, str(current_row))

        if len(self.__tables) != table_count:
            print("Listed tables in TOC does not match number of tabs in .xlsx file")
            self.__stop_report = True

    def format_sheet(self, sheet):
        self.adjust_dimensions(sheet)
        start_table_row = self.parse_col_names(sheet)
        self.unmerge_response_cols(sheet)
        if self.__has_stats:
            start_table_row = self.insert_names_row(sheet, start_table_row)
        else:
            self.create_col_names(sheet, start_table_row)
        if self.__is_numeric:
            self.insert_numeric_col(sheet, start_table_row)
        self.format_hyperlink(sheet)
        table = self.__tables.get(sheet.title)
        self.format_table_titles(sheet, table, start_table_row)
        self.format_banners(sheet, start_table_row)
        self.format_responses(sheet, start_table_row)
        self.format_table_contents(sheet, start_table_row)
        self.format_stats_def(sheet)
        self.add_table_borders(sheet)
        
    def adjust_dimensions(self, sheet):
        sheet.row_dimensions[1].height = self.__row_height

        sheet.row_dimensions[2].height = 36
        sheet.row_dimensions[3].height = 37
        sheet.column_dimensions["A"].width = 50
        sheet.column_dimensions["B"].width = 25

    def parse_col_names(self, sheet):
        print("Parsing column names")
        # calculate column names or determine end of table if applicable
        self.__col_names = []
        col_cell_row = None
        end_table_row = None

        for cell in sheet["A"]:
            if cell.value == "Column Names":
                col_cell_row = cell.row
                break

            if cell.value is not None:
                if "Total sample" in cell.value:
                    self.end_table_row = cell.row
                    break

                if "Filter:" in cell.value:
                    self.end_table_row = cell.row
                    break

        # pull column names
        if col_cell_row is not None:
            self.__has_stats = True
            for cell in sheet[col_cell_row]:
                if cell.value is not None and cell.value != "Column Names":
                    self.__col_names.append(cell.value)

            sheet.delete_rows(col_cell_row, 1)
            self.end_table_row = col_cell_row
        else:
            self.__has_stats = False
            self.end_table_row -= 1

        # determined where banners start
        col_index = 0
        banner_row = 4
        for column_name in self.__extend_alphabet:
            current_cell = "%s%s" % (column_name, str(banner_row))
            if sheet[current_cell].value is not None:
                banner_col = column_name
                break
            col_index += 1

        self.banner_col_index = col_index

        # determine where table details start
        start_table_row = 1
        for cell in sheet[banner_col]:
            if cell.number_format != "General":
                break
            else:
                start_table_row += 1

        self.__is_numeric = False
        # determine if table is numeric scale
        check_row = 1
        for cell in sheet[self.__extend_alphabet[self.banner_col_index+1]]:
            if cell.number_format == "0.0":
                self.__is_numeric = True
                break

        if self.__is_numeric:
            self.numeric_details = sheet["A3"].value.split("\n")
            for n, i in enumerate(self.numeric_details):
                if i == "Column n":
                    self.numeric_details[n] = "Population (n)"

        self.end_table_row += 1

        return start_table_row

    def unmerge_response_cols(self, sheet):
        print("Unmerge response columns")

        merged_ranges =[]  
        for group in sheet.merged_cells.ranges:
            merged_ranges.append(group)

        if self.__is_numeric:
            for range in merged_ranges:
                str_range = str(range)
                sheet.unmerge_cells(str_range)
        else:
            col_index = 0
            while col_index < self.banner_col_index:
                col_letter = self.__extend_alphabet[col_index]
                to_delete = []
                for range in merged_ranges:
                    str_range = str(range)
                    range_start_end = str_range.split(":")
                    start_col = re.match("([A-Z]+)\d+", range_start_end[0]).group(1)
                    end_col = re.match("([A-Z]+)\d+", range_start_end[1]).group(1)
                    if col_letter == start_col or col_letter == end_col:
                        to_delete.append(range)
                        sheet.unmerge_cells(str_range)
                for used_range in to_delete:
                    merged_ranges.remove(used_range)
                col_index += 1

    def insert_names_row(self, sheet, start_table_row):
        print("Insert names row")
        banner_row = start_table_row - 2
        top_banner_cell = "%s%s" % (self.__extend_alphabet[self.banner_col_index], str(banner_row))

        sheet.insert_rows(start_table_row - 1)
        sheet.row_dimensions[start_table_row - 1].height = 16

        current_col = self.banner_col_index
        for letter in self.__col_names:
            col_letter = self.__extend_alphabet[current_col]
            current_cell = "%s%s" % (col_letter, str(start_table_row - 1))
            sheet[current_cell].value = "[%s]" % letter
            sheet[current_cell].font = self.__font_reg
            sheet[current_cell].alignment = self.__align_center
            current_col += 1

        start_table_row += 1
        return start_table_row

    def create_col_names(self, sheet, start_table_row):
        banner_row = start_table_row - 2
        top_banner_cell = "%s%s" % (self.__extend_alphabet[self.banner_col_index], str(banner_row))

        # adjust list with the correct no of items for future calculations
        banner_row = start_table_row - 1
        current_col = self.banner_col_index
        current_cell = "%s%s" % (self.__extend_alphabet[current_col], str(banner_row))
        while sheet[current_cell].value is not None:
            self.__col_names.append(sheet[current_cell].value)
            current_col += 1
            current_cell = "%s%s" % (self.__extend_alphabet[current_col], str(banner_row))

    def insert_numeric_col(self, sheet, start_table_row):
        col_index = self.banner_col_index + 1
        sheet.insert_cols(col_index)
        self.banner_col_index += 1

        if self.__has_stats:
            banner_row = start_table_row - 3
        else:
            banner_row = start_table_row - 2

        # merge banner cells
        col_adjust = self.banner_col_index
        is_first = True
        merge_cells = []

        current_col = self.banner_col_index
        while current_col < len(self.__col_names) + col_adjust:
            current_cell = "%s%s" % (self.__extend_alphabet[current_col], str(banner_row))
            if sheet[current_cell].value is not None:
                if is_first:
                    is_first = False
                else:
                    range = "%s:%s" % (merge_cells[0], merge_cells[len(merge_cells)-1])
                    merge_cells = []
                    sheet.merge_cells(range)
                merge_cells.append(sheet[current_cell].coordinate)
            else:
                merge_cells.append(sheet[current_cell].coordinate)
            current_col += 1

        range = "%s:%s" % (merge_cells[0], merge_cells[len(merge_cells)-1])
        merge_cells = []
        sheet.merge_cells(range)

    def format_hyperlink(self, sheet):
        # check if we need to add a row
        if sheet["A1"].value != "Back to TOC":
            sheet.insert_rows(1)
        else:
            sheet["A1"].value = ""

        # back to TOC cell
        sheet["A3"].value = ""
        sheet["A1"].value = '=HYPERLINK("#TOC!A1","Return to Table of Contents")'
        sheet["A1"].font = self.__font_back_hyperlink
        sheet["A1"].alignment = self.__align_left

    def format_table_titles(self, sheet, table, start_table_row):
        print("Format table titles")
        current_row = 1
        current_col = 0

        col_adjust = self.banner_col_index
        current_col = 0
        for col in range(len(self.__col_names) + col_adjust):
            col_letter = self.__extend_alphabet[col]
            current_cell = "%s%s" % (col_letter, str(current_row))
            sheet[current_cell].fill = self.__header_fill
            next_row_cell = "%s%s" % (col_letter, str(current_row + 1))
            sheet[next_row_cell].fill = self.__table_fill
            current_col += 1

        col_adjust += 1

        # add logos
        current_cell = "%s%s" % (self.__extend_alphabet[current_col - col_adjust], str(current_row))
        logo = Image(self.__image_path)
        sheet.add_image(logo, current_cell)

        # merge base description cells
        table_title_row = 2
        base_desc_col_index = current_col-col_adjust
        start_range = "%s%s" % (self.__extend_alphabet[base_desc_col_index], str(table_title_row))
        end_range = "%s%s" % (self.__extend_alphabet[len(self.__col_names)+1], str(table_title_row))
        merge_range = "%s:%s" % (start_range, end_range)
        sheet.merge_cells(merge_range)
        
        # add base description
        desc_location = "%s%s" % (self.__extend_alphabet[base_desc_col_index], str(table_title_row))
        sheet[desc_location].font = self.__font_reg
        sheet[desc_location].alignment = self.__align_names
        sheet[desc_location] = "Base - %s" % table.description

        # merge table title row cells
        start_range = "%s%s" % ("A", str(table_title_row))
        end_range = "%s%s" % (self.__extend_alphabet[self.banner_col_index], str(table_title_row))
        merge_range = "%s:%s" % (start_range, end_range)
        sheet.merge_cells(merge_range)

        # add table title
        sheet["A2"].font = self.__font_reg
        sheet["A2"].alignment = self.__align_names
        table_title = sheet["A2"].value
        final_title = "%s - %s" % (sheet.title, table.prompt)
        sheet["A2"].value = final_title

        # merge top corner of table
        start_range = "A3"
        end_range = "%s%s" % (self.__extend_alphabet[self.banner_col_index-1],str(start_table_row-1))
        merge_range = "%s:%s" % (start_range, end_range)
        sheet.merge_cells(merge_range)

    def format_banners(self, sheet, start_row):
        print("format banners")
        col_adjust = self.banner_col_index
        if self.__has_stats is False:
            sheet.row_dimensions[4].height = 40

        current_row = start_row-1
        while current_row > 1:
            col_index = 1
            while col_index < len(self.__col_names) + col_adjust:
                current_cell = "%s%s" % (self.__extend_alphabet[col_index], str(current_row))
                sheet[current_cell].font = self.__font_reg
                sheet[current_cell].alignment = self.__align_banners
                col_index += 1
            current_row -= 1

    def format_responses(self, sheet, start_row):
        print("Format responses")
        if self.__is_numeric:
            self.format_numeric_details(sheet, start_row)
            response_col = self.banner_col_index - 2
        else:
            response_col = self.banner_col_index - 1

        self.__border_response_rows = set()
        merge_cells = []
        is_top = True

        current_row = start_row 
        end_row = self.end_table_row
        while current_row < end_row:
            current_cell = "%s%s" % (self.__extend_alphabet[response_col], str(current_row))
            if sheet[current_cell].value == "NET":
                sheet[current_cell].value = "Total"

            # merge response cells together
            if sheet[current_cell].value is not None:
                self.__border_response_rows.add(current_row)
                if is_top:
                    is_top = False
                else:
                    range = "%s:%s" % (merge_cells[0], merge_cells[len(merge_cells)-1])
                    merge_cells = []
                    sheet.merge_cells(range)
                merge_cells.append(sheet[current_cell].coordinate)
            else:
                merge_cells.append(sheet[current_cell].coordinate)

            sheet[current_cell].font = self.__font_bold
            sheet[current_cell].alignment = self.__align_left
            sheet[current_cell].fill = self.__table_fill
            current_row += 1

        range = "%s:%s" % (merge_cells[0], merge_cells[len(merge_cells)-1])
        merge_cells = []
        sheet.merge_cells(range)

        response_col -= 1
        while response_col >= 0:
            merge_cells = []
            is_top = True

            current_row = start_row 
            end_row = self.end_table_row
            while current_row < end_row:
                current_cell = "%s%s" % (self.__extend_alphabet[response_col], str(current_row))
                if sheet[current_cell].value == "NET":
                    sheet[current_cell].value = "Total"

                # merge response cells together
                if sheet[current_cell].value is not None:
                    self.__border_response_rows.add(current_row)
                    if is_top:
                        is_top = False
                    else:
                        range = "%s:%s" % (merge_cells[0], merge_cells[len(merge_cells)-1])
                        merge_cells = []
                        sheet.merge_cells(range)
                    merge_cells.append(sheet[current_cell].coordinate)
                else:
                    merge_cells.append(sheet[current_cell].coordinate)

                sheet[current_cell].font = self.__font_bold
                if response_col == 0:
                    sheet[current_cell].alignment = self.__align_top
                else:
                    sheet[current_cell].alignment = self.__align_left
                sheet[current_cell].fill = self.__table_fill
                current_row += 1
            range = "%s:%s" % (merge_cells[0], merge_cells[len(merge_cells)-1])
            merge_cells = []
            sheet.merge_cells(range)
            response_col -= 1

    def format_numeric_details(self, sheet, start_row):
        numeric_col = self.banner_col_index - 1
        sheet.column_dimensions[self.__extend_alphabet[numeric_col]].width = 15

        current_row = start_row
        end_row = self.end_table_row
        while current_row < end_row:
            for numeric_scale in self.numeric_details:
                num_cell = "%s%s" % (self.__extend_alphabet[numeric_col], str(current_row))
                sheet[num_cell].font = self.__font_bold
                sheet[num_cell].alignment = self.__align_left
                sheet[num_cell].fill = self.__table_fill
                sheet[num_cell].value = "%s" % numeric_scale

                current_row += 1

    def format_table_contents(self, sheet, start_row):
        print("format table")
        current_row = start_row
        col_adjust = self.banner_col_index

        # change font/alignment and highlight significant cells
        while current_row < self.end_table_row:
            col_no = self.banner_col_index
            while col_no < len(self.__col_names) + col_adjust:
                current_cell = "%s%s" % (self.__extend_alphabet[col_no], str(current_row))
                sheet[current_cell].font = self.__font_reg
                sheet[current_cell].alignment = self.__align_center
                if sheet[current_cell].data_type == 's':
                    value = sheet[current_cell].value.upper()
                    value_list = value.split(" ")
                    if len(value_list) > 1:
                        sheet[current_cell].fill = self.__hi_significant_fill
                    elif sheet[current_cell].value.isupper():
                        sheet[current_cell].fill = self.__hi_significant_fill
                    elif sheet[current_cell].value.islower():
                        sheet[current_cell].fill = self.__lo_significant_fill
                col_no += 1
            current_row += 1

    def format_stats_def(self, sheet):
        print("format stats def")
        current_row = self.end_table_row
        while True:
            current_cell = "A%s" % str(current_row)
            if sheet[current_cell].value is not None:
                sheet[current_cell].font = self.__font_small
            else:
                break
            current_row += 1

    def add_table_borders(self, sheet):
        print("table borders")
        col_adjust = self.banner_col_index
        for row_no in self.__border_response_rows:
            col_no = 0
            while col_no < len(self.__col_names) + col_adjust:
                current_cell = "%s%s" % (self.__extend_alphabet[col_no], str(row_no))
                sheet[current_cell].border = self.__thin_top
                col_no += 1

        col_no = 0
        while col_no < len(self.__col_names) + col_adjust:
            current_cell = "%s%s" % (self.__extend_alphabet[col_no], str(self.end_table_row))
            sheet[current_cell].border = self.__thick_top
            col_no += 1

        current_row = 1
        while current_row < self.end_table_row:
            current_cell = "%s%s" % (self.__extend_alphabet[col_no], str(current_row))
            sheet[current_cell].border = self.__thick_left
            current_row += 1

    def save(self, path_to_output):
        self.__workbook.save(path_to_output)

class TOCTable(object):

    def __init__(self, table_name):
        self.__name = str(table_name)

    @property
    def prompt(self):
        return self.__prompt

    @property
    def description(self):
        return self.__description

    @property
    def name(self):
        return self.__name

    @prompt.setter
    def prompt(self, prompt):
        self.__prompt = prompt

    @description.setter
    def description(self, desc):
        self.__description = desc
