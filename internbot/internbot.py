import base
import crosstabs
import topline
import rnc_automation
import Tkinter
import tkMessageBox
import tkFileDialog
import os, subprocess, platform
import csv
from collections import OrderedDict
from years_window import YearsWindow
from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image

class Internbot:

    def __init__ (self, root):
            self.__window = root
            self.main_buttons()
            self.fpath = os.path.join(os.path.expanduser("~"), "Desktop")
            self.__embedded_fields = []

    def main_buttons(self):
        btn_xtabs = Tkinter.Button(self.__window, text="Run crosstabs", command=self.software_tabs_menu, relief=Tkinter.GROOVE)
        btn_report = Tkinter.Button(self.__window, text="Run topline report", command=self.topline_menu, relief=Tkinter.GROOVE)
        btn_appen = Tkinter.Button(self.__window, text="Run topline appendix", command=self.append_menu, relief=Tkinter.GROOVE)
        btn_rnc = Tkinter.Button(self.__window, text="Run RNC", command=self.rnc_menu, relief=Tkinter.GROOVE)
        btn_quit = Tkinter.Button(self.__window, text="Quit", command=self.__window.destroy, relief=Tkinter.GROOVE)
        btn_xtabs.pack(padx=2, side=Tkinter.LEFT, expand=True)
        btn_report.pack(padx=2, side=Tkinter.LEFT, expand=True)
        btn_appen.pack(padx=2, side=Tkinter.LEFT, expand=True)
        btn_rnc.pack(padx=2, side=Tkinter.LEFT, expand=True)
        btn_quit.pack(padx=2, side=Tkinter.LEFT, expand=True)
        
        self.menubar = Tkinter.Menu(self.__window)
        menu_xtabs = Tkinter.Menu(self.menubar, tearoff = 0)
        menu_xtabs.add_command(label="Variable Script", command=self.variable_script)
        menu_xtabs.add_command(label="Table Script", command=self.table_script)
        menu_xtabs.add_command(label="Build Report", command=self.build_xtabs)
        self.menubar.add_cascade(label="Crosstabs", menu=menu_xtabs)
        
        menu_report = Tkinter.Menu(self.menubar, tearoff = 0)
        menu_report.add_command(label="Run Topline", command=self.topline_menu)
        menu_report.add_command(label="Run Appendix", command=self.append_menu)
        self.menubar.add_cascade(label="Topline", menu=menu_report)
        
        menu_rnc = Tkinter.Menu(self.menubar, tearoff=0)
        menu_rnc.add_command(label="Scores Topline", command=self.scores_window)
        menu_rnc.add_command(label="Issue Trended", command=self.issue_trended_window)
        menu_rnc.add_command(label="Trended Score", command=self.trended_scores_window)
        self.menubar.add_cascade(label="RNC", menu=menu_rnc)
        
        menu_quit = Tkinter.Menu(self.menubar, tearoff = 0)
        menu_quit.add_command(label="Good Bye", command=self.__window.destroy)
        self.menubar.add_cascade(label="Quit", menu=menu_quit)
        self.__window.config(menu=self.menubar)
        
    def software_tabs_menu(self):
        sft_window = Tkinter.Toplevel(self.__window)
        sft_window.withdraw()
        x = self.__window.winfo_x()
        y = self.__window.winfo_y()
        sft_window.geometry("300x200+%d+%d" % (x + 150, y + 100))
        message = "Please select crosstabs software to use"
        Tkinter.Label(sft_window, text = message).pack()
        btn_spss = Tkinter.Button(sft_window, text="SPSS", command=self.tabs_menu, height=1, width=15)
        btn_q = Tkinter.Button(sft_window, text="Q Research", command=self.qtab_menu, height=1, width=15)
        btn_cancel = Tkinter.Button(sft_window, text="Cancel", command=sft_window.destroy, height=1, width=15)
        btn_cancel.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        btn_q.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        btn_spss.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        sft_window.deiconify()

    def tabs_menu(self):
        redirect_window = Tkinter.Toplevel(self.__window)
        redirect_window.withdraw()
        x = self.__window.winfo_x()
        y = self.__window.winfo_y()
        redirect_window.geometry("300x200+%d+%d" % (x + 150, y + 100))
        message = "Please select the files to produce."
        Tkinter.Label(redirect_window, text = message).pack()
        btn_var = Tkinter.Button(redirect_window, text="Variable script", command=self.variable_script, height=1, width=15)
        btn_tab = Tkinter.Button(redirect_window, text="Table script", command=self.table_script, height=1, width=15)
        btn_compile = Tkinter.Button(redirect_window, text="Build report", command=self.build_xtabs, height=1, width=15)
        btn_cancel = Tkinter.Button(redirect_window, text="Cancel", command=redirect_window.destroy, height=1, width=15)
        btn_cancel.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        btn_compile.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        btn_tab.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        btn_var.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        redirect_window.deiconify()

    def qtab_menu(self):
        qtab_window = Tkinter.Toplevel(self.__window)
        qtab_window.withdraw()
        x = self.__window.winfo_x()
        y = self.__window.winfo_y()
        qtab_window.geometry("300x200+%d+%d" % (x + 150, y + 100))
        btn_build = Tkinter.Button(qtab_window, text="Format Report", command=self.qtab_build, height=1, width=15)
        btn_bases = Tkinter.Button(qtab_window, text="Propogate Bases", command=self.qtab_bases, height=1, width=15)
        btn_cancel = Tkinter.Button(qtab_window, text="Cancel", command=qtab_window.destroy, height=1, width=15)
        btn_cancel.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        btn_bases.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        btn_build.pack(padx=5, side=Tkinter.BOTTOM, expand=True)
        qtab_window.deiconify()

    def qtab_build(self):
        ask_tables = tkMessageBox.askokcancel("Select Q Research report file", "Please select the Q Research report file.")
        if ask_tables is True:
            reportfilename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select report file",filetypes = (("excel files","*.xlsx"),("all files","*.*")))
            if reportfilename is not "":
                parser = crosstabs.Format_Q_Report.QParser(reportfilename)
                ask_output = tkMessageBox.askokcancel("Select Output Directory", "Please select the directory for final report.")
                if ask_output is True: # user selected ok
                    savedirectory = tkFileDialog.askdirectory()
                    if savedirectory is not "":
                        self.tables = parser.format_report(savedirectory)
                        with open(savedirectory+"/tables.csv", "wb") as file:
                            writer = csv.writer(file, delimiter=",", quotechar= "\"")
                            writer.writerow(['Table No']+['Prompt']+['Base Description']+['Base Size'])
                            for key, value in self.tables.iteritems():
                                prompt = "%s" % value.prompt
                                writer.writerow([value.name]+[prompt])



    def qtab_bases(self):
        ask_tables = tkMessageBox.askokcancel("Select Q Research report file", "Please select the Q Research report file.")
        if ask_tables is True:
            reportfilename =  "/Users/tristanbowler/Desktop/tables test.csv"
            if reportfilename is not "":
                with open(reportfilename) as tables_file:
                    reader = csv.DictReader(tables_file, delimiter=",")

                    for row in reader:
                        index = int(row['Table No'])
                        self.tables[index].description = row['Base Description']
                        self.tables[index].size = row['Base Size']

                    reportfilename = "/Users/tristanbowler/Desktop/Crosstab Report.xlsx"
                    workbook = load_workbook(reportfilename)
                    current_row = 3
                    for key, table in self.tables.iteritems():
                        if int(table.name) < 10:
                            sheet_name = "Table 0%s" % str(table.name)
                        else:
                            sheet_name = "Table %s" % str(table.name)
                        sheet = workbook.get_sheet_by_name(sheet_name)
                        sheet[table.location].value = table.description
                        toc_sheet = workbook.get_sheet_by_name("TOC")
                        base_description_cell = "C%s" % str(current_row)
                        base_size_cell = "D%s" % str(current_row)
                        toc_sheet[base_description_cell].value = table.description
                        toc_sheet[base_size_cell].value = table.size
                        current_row += 1

                    workbook.save("/Users/tristanbowler/Desktop/Final Report.xlsx")

   






    def topline_menu(self):
        self.redirect_window = Tkinter.Toplevel(self.__window)
        self.redirect_window.withdraw()
        x = self.__window.winfo_x()
        y = self.__window.winfo_y()
        self.redirect_window.geometry("250x100+%d+%d" % (x + 175, y + 150))
        self.redirect_window.title("Y2 Topline Report Automation")
        message = "Please open a survey file."
        Tkinter.Label(self.redirect_window, text = message).pack(expand=True)

        # round details
        round_frame = Tkinter.Frame(self.redirect_window)
        round_frame.pack(side = Tkinter.TOP, expand = True)

        round_label = Tkinter.Label(round_frame, text="Round number:")
        round_label.pack(side = Tkinter.LEFT, expand=True)
        self.round_entry = Tkinter.Entry(round_frame)
        self.round_entry.pack(side=Tkinter.RIGHT, expand=True)
        
        btn_open = Tkinter.Button(self.redirect_window, text = "Open", command = self.read_topline)
        btn_cancel = Tkinter.Button(self.redirect_window, text = "Cancel", command = self.redirect_window.destroy)
        btn_open.pack(ipadx = 10, side = Tkinter.LEFT, expand=True)
        btn_cancel.pack(ipadx = 10, side = Tkinter.LEFT, expand=True)
        self.redirect_window.deiconify()

    def append_menu(self):
        self.redirect_window = Tkinter.Toplevel(self.__window)
        self.redirect_window.withdraw()
        x = self.__window.winfo_x()
        y = self.__window.winfo_y()
        self.redirect_window.geometry("250x100+%d+%d" % (x + 175, y + 150))
        self.redirect_window.title("Y2 Topline Appendix Report Automation")
        message = "Please open a appendix file."
        Tkinter.Label(self.redirect_window, text = message).pack(expand=True)

        btn_open = Tkinter.Button(self.redirect_window, text = "Open", command = self.read_append)
        btn_cancel = Tkinter.Button(self.redirect_window, text = "Cancel", command = self.redirect_window.destroy)
        btn_open.pack(ipadx = 10, side = Tkinter.LEFT, expand=True)
        btn_cancel.pack(ipadx = 10, side = Tkinter.LEFT, expand=True)
        self.redirect_window.deiconify()

    def rnc_menu(self):
        self.redirect_window = Tkinter.Toplevel(self.__window)
        self.redirect_window.withdraw()
        x = self.__window.winfo_x()
        y = self.__window.winfo_y()
        self.redirect_window.geometry("250x150+%d+%d" % (x + 175, y + 125))
        self.redirect_window.title("RNC Scores Topline Automation")
        message = "Please open a model scores file."
        Tkinter.Label(self.redirect_window, text = message).pack(expand=True)
        btn_topline = Tkinter.Button(self.redirect_window, text="Scores Topline Report", command=self.scores_window, height = 1, width = 20)
        btn_trended = Tkinter.Button(self.redirect_window, text="Issue Trended Report", command=self.issue_trended_window, height = 1, width = 20)
        btn_ind_trended = Tkinter.Button(self.redirect_window, text="Trended Score Reports", command=self.trended_scores_window, height = 1, width = 20)
        btn_cancel = Tkinter.Button(self.redirect_window, text="Cancel", command = self.redirect_window.destroy, height = 1, width = 20)
        btn_cancel.pack(ipadx = 5, side = Tkinter.BOTTOM, expand=False)
        btn_topline.pack(ipadx=5, side=Tkinter.BOTTOM, expand=False)
        btn_trended.pack(ipadx=5, side = Tkinter.BOTTOM, expand=False)
        btn_ind_trended.pack(ipadx = 5, side = Tkinter.BOTTOM, expand=False)
        self.redirect_window.deiconify()

    def read_append(self):
        qsffilename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select survey file", filetypes = (("Qualtrics files","*.qsf"), ("all files", "*.*")))
        compiler = base.QSFSurveyCompiler()
        survey = compiler.compile(qsffilename)            
        filename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select open ends file",filetypes = (("comma seperated files","*.csv"),("all files","*.*")))
        output = tkFileDialog.askdirectory()
        builder = topline.QSF.ReportGenerator(survey)
        builder.generate_appendix("appendix_template.docx", filename, output)
        self.redirect_window.destroy()

    def read_topline(self):
		try:
			filename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select survey file",filetypes = (("Qualtrics files","*.qsf"),("comma seperated files","*.csv"),("all files","*.*")))
			if filename is not "":
				self.isQSF = False
				if ".qsf" in filename:
					pass
				elif ".csv" in filename:
					round_int = self.round_entry.get()
					is_trended = False
					if round_int == "":
						round_int = 1
					else:
						round_int = int(round_int)
						if round_int != 1:
							is_trended = True
					self.report = topline.CSV.ReportGenerator(filename, round_int)
					self.redirect_window.destroy()
					self.round = round_int
					if is_trended is True:
						self.year_window_setup()
					else:
						self.build_topline_report(self.isQSF, self.report)
		except Exception as e:
				tkMessageBox.showerror("Error", "An error occurred\n"+ str(e))

    def year_window_setup(self):
        try:
            self.year_window = Tkinter.Toplevel(self.__window)
            self.year_window.withdraw()
            x = self.__window.winfo_x()
            y = self.__window.winfo_y()
            self.year_window.geometry("500x%d+%d+%d" % (100+self.round*25, x + 50, y + (200-(100+self.round*25)/2)))
            self.year_window.title("Trended report years")
            message = "Please input the applicable years for the trended topline report."
            Tkinter.Label(self.year_window, text=message).pack(expand=True)

            year_frame = Tkinter.Frame(self.year_window)
            year_frame.pack(side = Tkinter.TOP, expand = True)
            self.year_window_obj = YearsWindow(self.__window, self.year_window, self.round)
            self.year_window_obj.packing_years(year_frame)
            btn_finish = Tkinter.Button(self.year_window, text = "Done", command=self.build_topline_leadup, height = 1, width = 20)
            btn_cancel = Tkinter.Button(self.year_window, text = "Cancel", command=self.year_window.destroy, height = 1, width = 20)
            btn_finish.pack(ipadx=5, side = Tkinter.LEFT, expand=False)
            btn_cancel.pack(ipadx=5, side = Tkinter.RIGHT, expand=False)
            self.year_window.deiconify()
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n"+ str(e))

    def build_topline_leadup(self):
        years = self.year_window_obj.read_years()
        self.build_topline_report(self.isQSF, self.report, years)

    def build_topline_report(self, isQSF, report, years=[]):
        try:
            template_file = open("topline_template.docx", "r")
            ask_output = tkMessageBox.askokcancel("Output directory", "Please select the directory for finished report.")
            if ask_output is True:
                savedirectory = tkFileDialog.askdirectory()
                if savedirectory is not "":
                    if isQSF is True:
                        pass
                    else:
                        report.generate_topline(template_file, savedirectory, years)
                        open_files = tkMessageBox.askyesno("Info", "Done!\nWould you like to open your finished files?")
                        if open_files is True:
                            self.open_file_for_user(savedirectory+"/topline_report.docx")
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def variable_script(self):
        try:
            ask_qsf = tkMessageBox.askokcancel("Select Qualtrics File", "Please select the Qualtrics survey .qsf file.")
            if ask_qsf is True: # user selected ok
                qsffilename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select Qualtrics survey file",filetypes = (("Qualtrics file","*.qsf"),("all files","*.*")))
                if qsffilename is not "":
                    compiler = base.QSFSurveyCompiler()
                    survey = compiler.compile(qsffilename)
                    ask_output = tkMessageBox.askokcancel("Output directory", "Please select the directory for finished variable script.")
                    if ask_output is True: # user selected ok
                        savedirectory = tkFileDialog.askdirectory()
                        if savedirectory is not "":
                            variables = crosstabs.Format_SPSS_Report.Generate_Prelim_SPSS_Script.SPSSTranslator()
                            tables = crosstabs.Format_SPSS_Report.Generate_Prelim_SPSS_Script.TableDefiner()
                            variables.define_variables(survey, savedirectory)
                            tables.define_tables(survey, savedirectory)
                            open_files = tkMessageBox.askyesno("Info", "Done!\nWould you like to open your finished files?")
                            if open_files is True:
                                self.open_file_for_user(savedirectory+"/Tables to run.csv")
                                self.open_file_for_user(savedirectory+"/rename variables.sps")
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n"+ str(e))

    def table_script(self):
        try:
            ask_tables = tkMessageBox.askokcancel("Select Tables to Run.csv File", "Please select the tables to run .csv file.")
            if ask_tables is True:
                self.tablesfilename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select tables file",filetypes = (("comma seperated files","*.csv"),("all files","*.*")))
                if self.tablesfilename is not "":
                    ask_banners = tkMessageBox.askokcancel("Banner selection", "Please insert/select the banners for this report.")
                    if ask_banners is True:
                        names = crosstabs.Format_SPSS_Report.Generate_Table_Script.TablesParser().pull_table_names(self.tablesfilename)
                        titles = crosstabs.Format_SPSS_Report.Generate_Table_Script.TablesParser().pull_table_titles(self.tablesfilename)
                        bases = crosstabs.Format_SPSS_Report.Generate_Table_Script.TablesParser().pull_table_bases(self.tablesfilename)
                        self.banner_window(names, titles, bases)
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n"+ str(e))
            
   
    def banner_window(self, names, titles, bases):
        try:
            self.edit_window = Tkinter.Toplevel(self.__window)
            self.edit_window.withdraw()
            x = self.__window.winfo_x()
            y = self.__window.winfo_y()
            self.edit_window.geometry("1500x500+%d+%d" % (x - 450 , y - 50))

            self.edit_window.title("Banner selection")
            
            self.boxes_frame = Tkinter.Frame(self.edit_window)
            self.boxes_frame.pack(side=Tkinter.LEFT, fill=Tkinter.BOTH)
            horiz_scrollbar = Tkinter.Scrollbar(self.boxes_frame)
            horiz_scrollbar.config(orient= Tkinter.HORIZONTAL )
            horiz_scrollbar.pack(side=Tkinter.BOTTOM, fill=Tkinter.X)
            vert_scrollbar = Tkinter.Scrollbar(self.boxes_frame)
            vert_scrollbar.pack(side=Tkinter.RIGHT, fill=Tkinter.Y)
            
            
            self.tables_box = Tkinter.Listbox(self.boxes_frame, selectmode="multiple", width=80, height=15, yscrollcommand=vert_scrollbar.set, xscrollcommand = horiz_scrollbar.set)
            self.tables_box.pack(padx = 15, pady=10,expand=True, side = Tkinter.LEFT, fill=Tkinter.BOTH)
            self.banners_box = Tkinter.Listbox(self.edit_window)
            self.banners_box.pack(padx = 15, pady=10, expand=True, side=Tkinter.RIGHT, fill=Tkinter.BOTH)
            
            

            index = 0
            while index < len(names):
                self.tables_box.insert(Tkinter.END, names[index] + ": " + titles[index])
                index += 1
            
            vert_scrollbar.config(command=self.tables_box.yview)
            horiz_scrollbar.config(command=self.tables_box.xview)
            
            buttons_frame = Tkinter.Frame(self.edit_window)
            buttons_frame.pack(side = Tkinter.RIGHT, fill=Tkinter.BOTH)
            btn_up = Tkinter.Button(buttons_frame, text = "Up", command = self.shift_up)
            btn_down = Tkinter.Button(buttons_frame, text = "Down", command = self.shift_down)
            btn_insert = Tkinter.Button(buttons_frame, text = "Insert", command = self.insert_banner)
            btn_edit = Tkinter.Button(buttons_frame, text =   "Edit", command = self.parse_selection)
            btn_create = Tkinter.Button(buttons_frame, text = "Create", command = self.create_banner)
            btn_remove = Tkinter.Button(buttons_frame, text = "Remove", command = self.remove_banner)
            btn_done = Tkinter.Button(buttons_frame, text = "Done", command = self.finish_banner)

            btn_done.pack(side=Tkinter.BOTTOM, pady=15)
            btn_remove.pack(side=Tkinter.BOTTOM)
            btn_create.pack(side=Tkinter.BOTTOM)
            btn_edit.pack(side=Tkinter.BOTTOM)
            btn_insert.pack(side=Tkinter.BOTTOM, pady=5)
            btn_down.pack(side=Tkinter.BOTTOM)
            btn_up.pack(side=Tkinter.BOTTOM) 

            self.edit_window.deiconify()
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred" + str(e))

    def shift_up(self):
        current_banners = []
        for banner in self.banners_box.curselection():
            current_banners.append(banner)

        if len(current_banners) > 0:
            self.shift_banners_up()

        current_tables = []
        for table in self.tables_box.curselection():
            current_tables.append(table)

        if len(current_tables) > 0:
            self.shift_tables_up(current_tables)

    def shift_banners_up(self):
        old_index = -1
        new_index = -1
        for index in self.banners_box.curselection():
            table = self.banners_box.get(int(index))
            old_index = index
            new_index = old_index - 1
            if old_index > -1 and new_index > -1:
                self.banners_box.delete(old_index)
                self.banners_box.insert(new_index, table)
                self.banners_box.selection_clear(old_index, old_index)
                self.banners_box.selection_set(new_index)

    def shift_tables_up(self, current_tables):
        shifted_indexes = []
        index = 0
        while index < len(current_tables):
            selection_index = current_tables[index]
            new_index = selection_index - 1
            if new_index > -1:
                if new_index not in shifted_indexes:
                    shifted_indexes.append(new_index)
                else:
                    shifted_indexes.append(selection_index)
            else:
                shifted_indexes.append(selection_index)
            index += 1

        iteration = 0
        self.tables_box.selection_clear(0, Tkinter.END)
        for table_index in current_tables:
            is_highlighted = False
            current_table = self.tables_box.get(table_index)
            if '#C00201' in self.tables_box.itemconfig(table_index)['foreground']:
                is_highlighted = True
            self.tables_box.delete(table_index)
            self.tables_box.insert(shifted_indexes[iteration], current_table)
            if is_highlighted:
                self.tables_box.itemconfig(shifted_indexes[iteration], {'fg': '#C00201'})
            self.tables_box.selection_set(shifted_indexes[iteration])
            iteration += 1

    def shift_down(self):
        current_banners = []
        for banner in self.banners_box.curselection():
            current_banners.append(banner)

        if len(current_banners) > 0:
            self.shift_banners_down()

        current_tables = []
        for table in self.tables_box.curselection():
            current_tables.append(table)

        if len(current_tables) > 0:
            self.shift_tables_down(current_tables)

    def shift_banners_down(self):
        old_index = -1
        new_index = -1
        for index in self.banners_box.curselection():
            table = self.banners_box.get(int(index))
            old_index = index
            new_index = old_index + 1
            if new_index < self.banners_box.size():
                self.banners_box.delete(old_index)
                self.banners_box.insert(new_index, table)
                self.banners_box.selection_clear(old_index, old_index)
                self.banners_box.selection_set(new_index)

    def shift_tables_down(self, current_tables):
        shifted_indexes = []
        index = len(current_tables) - 1
        while index >= 0:
            selection_index = current_tables[index]
            new_index = selection_index + 1
            if new_index < self.tables_box.size():
                if new_index not in shifted_indexes:
                    shifted_indexes.append(new_index)
                else:
                    shifted_indexes.append(selection_index)
            else:
                shifted_indexes.append(selection_index)
            index -= 1

        current_tables.sort(reverse=True)
        shifted_indexes.sort(reverse=True)

        iteration = 0
        self.tables_box.selection_clear(0, Tkinter.END)
        for table_index in current_tables:
            is_highlighted = False
            current_table = self.tables_box.get(table_index)
            if '#C00201' in self.tables_box.itemconfig(table_index)['foreground']:
                is_highlighted = True
            self.tables_box.delete(table_index)
            self.tables_box.insert(shifted_indexes[iteration], current_table)
            if is_highlighted:
                self.tables_box.itemconfig(shifted_indexes[iteration], {'fg': '#C00201'})
            self.tables_box.selection_set(shifted_indexes[iteration])
            iteration += 1

    def insert_banner(self):
        indexes_to_clear = []
        for index in self.tables_box.curselection():
            table = self.tables_box.get(int(index))
            question = table.split(": ")
            name = question[0]
            title = question[1]
            self.banners_box.insert(Tkinter.END, name + ": " + title)
            self.tables_box.itemconfig(index, {'fg': '#C00201'})
            self.tables_box.selection_clear(index, index)

    def create_banner(self, initial_name='', intial_title=''):
        create_window = Tkinter.Toplevel(self.edit_window)
        create_window.title("Create banner")
        lbl_name = Tkinter.Label(create_window, text="Banner name:")
        lbl_title = Tkinter.Label(create_window, text="Banner title:")
        entry_name = Tkinter.Entry(create_window)
        entry_name.insert(0, initial_name)
        entry_title = Tkinter.Entry(create_window)
        entry_title.insert(0, intial_title)
        def create():
            name = entry_name.get()
            title = entry_title.get()
            self.banners_box.insert(Tkinter.END, name + ": " + title)
            self.__embedded_fields.append(name)
            self.tables_box.insert(0, name + ": " + title)
            self.tables_box.itemconfig(0, {'fg': '#C00201'})
            create_window.destroy()

        createButton = Tkinter.Button(create_window, text="Create", command=create)
        cancelButton = Tkinter.Button(create_window, text="Cancel", command=create_window.destroy)

        lbl_name.grid(row=0, column=0)
        lbl_title.grid(row=0, column=1)
        entry_name.grid(row=1, column=0)
        entry_title.grid(row=1, column=1)
        createButton.grid(row=2, column=0)
        cancelButton.grid(row=2, column=1)

    def parse_selection(self):
        name = ''
        title = ''
        for index in self.tables_box.curselection():
            table = self.tables_box.get(int(index))
            question = table.split(": ")
            name = question[0]
            title = question[1]
        if name is not '':
            self.edit_table(name, title, index)
        else:
            for index in self.banners_box.curselection():
                banner = self.banners_box.get(int(index))
                question = banner.split(": ")
                name = question[0]
                title = question[1]
                self.edit_banner(name, title, index)

    def edit_table(self, initial_names='', initial_title='', index=-1):
        edit_window = Tkinter.Toplevel(self.edit_window)
        edit_window.title("Edit banner")
        lbl_banner = Tkinter.Label(edit_window, text="Banner name:")
        entry_banner = Tkinter.Entry(edit_window)
        entry_banner.insert(0, initial_names)
        lbl_title = Tkinter.Label(edit_window, text="Title:")
        entry_title = Tkinter.Entry(edit_window)
        entry_title.insert(0, initial_title)

        def edit():
            banner_name = entry_banner.get()
            title_name = entry_title.get()
            self.tables_box.delete(int(index))
            self.tables_box.insert(int(index), banner_name + ": " + title_name)
            edit_window.destroy()

        btn_cancel = Tkinter.Button(edit_window, text="Cancel", command=edit_window.destroy)
        btn_edit = Tkinter.Button(edit_window, text="Edit", command=edit)
        lbl_banner.grid(row = 0, column = 0)
        lbl_title.grid(row = 0, column = 1)
        entry_banner.grid(row = 1, column = 0)
        entry_title.grid(row = 1, column = 1)

        btn_edit.grid(row = 2, column = 0)
        btn_cancel.grid(row = 2, column = 1)

    def edit_banner(self, initial_names='', initial_title='', index=-1):
        edit_window = Tkinter.Toplevel(self.edit_window)
        edit_window.title("Edit banner")
        lbl_banner = Tkinter.Label(edit_window, text="Banner name:")
        entry_banner = Tkinter.Entry(edit_window)
        entry_banner.insert(0, initial_names)
        lbl_title = Tkinter.Label(edit_window, text="Title:")
        entry_title = Tkinter.Entry(edit_window)
        entry_title.insert(0, initial_title)

        def edit():
            banner_name = entry_banner.get()
            title_name = entry_title.get()
            self.banners_box.delete(int(index))
            self.banners_box.insert(int(index), banner_name + ": " + title_name)
            edit_window.destroy()

        btn_cancel = Tkinter.Button(edit_window, text="Cancel", command=edit_window.destroy)
        btn_edit = Tkinter.Button(edit_window, text="Edit", command=edit)
        lbl_banner.grid(row = 0, column = 0)
        lbl_title.grid(row = 0, column = 1)
        entry_banner.grid(row = 1, column = 0)
        entry_title.grid(row = 1, column = 1)

        btn_edit.grid(row = 2, column = 0)
        btn_cancel.grid(row = 2, column = 1)

    def remove_banner(self):
        indexes_to_delete = []
        for index in self.tables_box.curselection():
            if '#C00201' not in self.tables_box.itemconfig(index)['foreground']:
                item = self.tables_box.get(index)
                indexes_to_delete.append(index)
                question = item.split(": ")
        indexes_to_delete.sort(reverse=True)
        for index in indexes_to_delete:
            self.tables_box.delete(int(index))
        for index in self.banners_box.curselection():
            item = self.banners_box.get(int(index))
            self.banners_box.delete(int(index))
            banner_to_remove = item.split(": ")
            index = 0
            while index < self.tables_box.size():
                to_compare = self.tables_box.get(index).split(": ")
                if to_compare[0] == banner_to_remove[0]:
                    self.tables_box.itemconfig(index, {'fg': '#000000'})
                index += 1

    def finish_banner(self):
        try:
            self.variable_entry = None
            ask_trended = tkMessageBox.askyesno("Trended Follow-up", "Is this a trended report?")
            if ask_trended is True:
                self.filter_variable_window()
            else:
                self.save_table_script()
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def filter_variable_window(self):
        self.create_window = Tkinter.Toplevel(self.__window)
        self.create_window.withdraw()
        x = self.__window.winfo_x()
        y = self.__window.winfo_y()
        self.create_window.geometry("250x100+%d+%d" % (x + 175, y + 150))
        self.create_window.title("Trended report details")

        # filtering  variable
        variable_frame = Tkinter.Frame(self.create_window)
        variable_frame.pack(side = Tkinter.TOP, expand=True)

        variable_frame = Tkinter.Label(variable_frame, text="Trended variable:")
        variable_frame.pack (side = Tkinter.LEFT, expand=True)
        self.variable_entry = Tkinter.Entry(variable_frame)
        self.variable_entry.pack(side=Tkinter.RIGHT, expand=True)

        # done and cancel buttons
        button_frame = Tkinter.Frame(self.create_window)
        button_frame.pack(side = Tkinter.TOP, expand=True)

        btn_cancel = Tkinter.Button(self.create_window, text = "Cancel", command = self.create_window.destroy)
        btn_cancel.pack(side = Tkinter.RIGHT, expand=True)
        btn_done = Tkinter.Button(self.create_window, text = "Done", command = self.save_table_script)
        btn_done.pack(side = Tkinter.RIGHT, expand=True)
        self.create_window.deiconify()

    def save_table_script(self):
        try:
            generator = crosstabs.Format_SPSS_Report.Generate_Table_Script.TableScriptGenerator()
            table_order = OrderedDict()
            banner_list = OrderedDict()
            for item in list(self.tables_box.get(0, Tkinter.END)):
                question = item.split(": ")
                table_order[question[0]] = question[1]
            for item in list(self.banners_box.get(0, Tkinter.END)):
                question = item.split(": ")
                banner_list[question[0]] = question[1]
            self.edit_window.destroy()
            ask_output = tkMessageBox.askokcancel("Output directory", "Please select the directory for finished table script.")
            filtering_variable = None
            if self.variable_entry is not None:
                variable = str(self.variable_entry.get())
                if variable != "":
                    filtering_variable = variable
                self.create_window.destroy()
            if ask_output is True:
                savedirectory = tkFileDialog.askdirectory()
                if savedirectory is not "":
                     generator.compile_scripts(self.tablesfilename, banner_list.keys(), self.__embedded_fields, filtering_variable, savedirectory)
                     open_files = tkMessageBox.askyesno("Info","Done!\nWould you like to open your finished files?")
                     if open_files is True:
                     	self.open_file_for_user(savedirectory + "/table script.sps")
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def build_xtabs(self):
        try:
            ask_directory = tkMessageBox.askokcancel("Select Tables Folder", "Please select the folder containing SPSS generated .xlsx table files.")
            if ask_directory is True:
                tablesdirectory = tkFileDialog.askdirectory()
                builder = crosstabs.Format_SPSS_Report.Parse_SPSS_Tables.CrosstabGenerator(tablesdirectory)
                ask_output = tkMessageBox.askokcancel("Output directory", "Please select the directory for finished report.")
                if ask_output is True:
                    outputdirectory = tkFileDialog.askdirectory()
                    if outputdirectory is not "":
                        builder.write_report(outputdirectory)
                        open_files = tkMessageBox.askyesno("Info", "Done!\nWould you like to open your finished files?")
                        if open_files is True:
                            self.open_file_for_user(outputdirectory + "/Crosstab Report.xlsx")
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def scores_window(self):
        try:
            okay = tkMessageBox.askokcancel("Select", "Select a model file")
            if okay is True:
                self.filename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select model file", filetypes = (("comma seperated files","*.csv"),("all files","*.*")))
                if self.filename is not "":
                    self.create_window = Tkinter.Toplevel(self.redirect_window)
                    self.create_window.withdraw()
                    x = self.__window.winfo_x()
                    y = self.__window.winfo_y()
                    self.create_window.geometry("250x100+%d+%d" % (x + 175, y + 150))
                    self.create_window.title("Scores Topline Report Details")
                    # location details
                    location_frame = Tkinter.Frame(self.create_window)
                    location_frame.pack(side = Tkinter.TOP, expand=True)

                    location_label = Tkinter.Label(location_frame, text="Reporting region:")
                    location_label.pack (side = Tkinter.LEFT, expand=True)
                    self.location_entry = Tkinter.Entry(location_frame)
                    self.location_entry.pack(side=Tkinter.RIGHT, expand=True)

                    # round details
                    round_frame = Tkinter.Frame(self.create_window)
                    round_frame.pack(side = Tkinter.TOP, expand=True)

                    round_label = Tkinter.Label(round_frame, text="Round number:")
                    round_label.pack(padx = 5, side = Tkinter.LEFT, expand=True)
                    self.round_entry = Tkinter.Entry(round_frame)
                    self.round_entry.pack(padx = 7, side=Tkinter.RIGHT, expand=True)

                    # done and cancel buttons
                    button_frame = Tkinter.Frame(self.create_window)
                    button_frame.pack(side = Tkinter.TOP, expand=True)

                    btn_cancel = Tkinter.Button(self.create_window, text = "Cancel", command = self.create_window.destroy)
                    btn_cancel.pack(side = Tkinter.RIGHT, expand=True)
                    btn_done = Tkinter.Button(self.create_window, text = "Done", command = self.scores_topline)
                    btn_done.pack(side = Tkinter.RIGHT, expand=True)
                    self.create_window.deiconify()
        except Exception as e:
                tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def scores_topline(self):
        try:
            filename = self.filename
            #fields entered by user
            report_location = self.location_entry.get()
            round = self.round_entry.get()
            if round is not "" and report_location is not "":
                self.create_window.destroy()
                report = rnc_automation.ScoresToplineReportGenerator(filename, int(round))
                ask_output = tkMessageBox.askokcancel("Output directory", "Please select the directory for finished report.")
                if ask_output is True:
                    savedirectory = tkFileDialog.askdirectory()
                    if savedirectory is not "":
                        report.generate_scores_topline(savedirectory, report_location, round)
                        open_files = tkMessageBox.askyesno("Info", "Done!\nWould you like to open your finished files?")
                        if open_files is True:
                            self.open_file_for_user(savedirectory + "/scores_topline.xlsx")
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def issue_trended_window(self):
        try:
            okay = tkMessageBox.askokcancel("Select", "Select a model file")
            if okay is True:
                self.filename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select model file", filetypes = (("comma seperated files","*.csv"),("all files","*.*")))
                if self.filename is not "":
                    self.create_window = Tkinter.Toplevel(self.redirect_window)
                    self.create_window.withdraw()
                    x = self.__window.winfo_x()
                    y = self.__window.winfo_y()
                    self.create_window.geometry("250x100+%d+%d" % (x + 175, y + 150))
                    self.create_window.title("Issue Trended Report Details")

                    # round details
                    round_frame = Tkinter.Frame(self.create_window)
                    round_frame.pack(side = Tkinter.TOP, expand=True)

                    round_label = Tkinter.Label(round_frame, text="Round number:")
                    round_label.pack(padx = 5, side = Tkinter.LEFT, expand=True)
                    self.round_entry = Tkinter.Entry(round_frame)
                    self.round_entry.pack(padx = 7, side=Tkinter.RIGHT, expand=True)

                    # done and cancel buttons
                    button_frame = Tkinter.Frame(self.create_window)
                    button_frame.pack(side = Tkinter.TOP, expand=True)

                    btn_cancel = Tkinter.Button(self.create_window, text = "Cancel", command = self.create_window.destroy)
                    btn_cancel.pack(side = Tkinter.RIGHT, expand=True)
                    btn_done = Tkinter.Button(self.create_window, text = "Done", command = self.issue_trended)
                    btn_done.pack(side = Tkinter.RIGHT, expand=True)
                    self.create_window.deiconify()
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def issue_trended(self):
        try:
            filename = self.filename
            #fields entered by the user
            round = self.round_entry.get()
            self.create_window.destroy()
            report = rnc_automation.IssueTrendedReportGenerator(filename, int(round))
            ask_output = tkMessageBox.askokcancel("Output directory", "Please select the directory for finished report.")
            if ask_output is True:
                savedirectory = tkFileDialog.askdirectory()
                if savedirectory is not "":
                    report.generate_issue_trended(savedirectory, round)
                    self.create_window.destroy
                    open_files = tkMessageBox.askyesno("Info", "Done!\nWould you like to open your finished files?")
                    if open_files is True:
                        self.open_file_for_user(savedirectory + "/trended.xlsx")
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))


    def trended_scores_window(self):
        try:
            okay = tkMessageBox.askokcancel("Select", "Select a model file")
            if okay is True:
                self.filename = tkFileDialog.askopenfilename(initialdir = self.fpath, title = "Select model file", filetypes = (("comma seperated files","*.csv"),("all files","*.*")))
                if self.filename is not "":
                    self.create_window = Tkinter.Toplevel(self.redirect_window)
                    self.create_window.withdraw()
                    x = self.__window.winfo_x()
                    y = self.__window.winfo_y()
                    self.create_window.geometry("250x100+%d+%d" % (x + 175, y + 150))
                    self.create_window.title("Trended Issue Reports Details")

                    # round details
                    round_frame = Tkinter.Frame(self.create_window)
                    round_frame.pack(side = Tkinter.TOP, expand=True)

                    round_label = Tkinter.Label(round_frame, text="Round number:")
                    round_label.pack(padx = 5, side = Tkinter.TOP, expand=True)
                    self.round_entry = Tkinter.Entry(round_frame)
                    self.round_entry.pack(padx = 7, side=Tkinter.BOTTOM, expand=True)

                    # done and cancel buttons
                    button_frame = Tkinter.Frame(self.create_window)
                    button_frame.pack(side = Tkinter.TOP, expand=True)

                    btn_cancel = Tkinter.Button(self.create_window, text = "Cancel", command = self.create_window.destroy)
                    btn_cancel.pack(side = Tkinter.RIGHT, expand=True)
                    btn_done = Tkinter.Button(self.create_window, text = "Done", command = self.trended_scores)
                    btn_done.pack(side = Tkinter.RIGHT, expand=True)

                    self.create_window.deiconify()
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def trended_scores(self):
        try:
            filename = self.filename
            #Field entered by user
            round = self.round_entry.get()
            if round is not "":
                self.create_window.destroy()
                report = rnc_automation.TrendedScoresReportGenerator(filename, int(round))
                ask_output = tkMessageBox.askokcancel("Output directory", "Please select the directory for finished report.")
                if ask_output is True:
                    savedirectory = tkFileDialog.askdirectory()
                    if savedirectory is not "":
                        report.generate_trended_scores(savedirectory, round)
                        self.create_window.destroy()
                        tkMessageBox.showinfo("Info", "Done!")
        except Exception as e:
            tkMessageBox.showerror("Error", "An error occurred\n" + str(e))

    def open_file_for_user(self, file_path):
        try:
            if os.path.exists(file_path):
                if platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', file_path))
                elif platform.system() == 'Windows':  # Windows
                    os.startfile(file_path)
            else:
                tkMessageBox.showerror("Error", "Error: Could not open file for you \n"+file_path)
        except IOError:
            tkMessageBox.showerror("Error", "Error: Could not open file for you \n" + file_path)

window = Tkinter.Tk()
window.withdraw()
window.title("Internbot: 01011001 00000010") # Internbot: Y2
if platform.system() == 'Windows':  # Windows
    window.iconbitmap('y2.ico')
screen_width = window.winfo_screenwidth()

screen_height = window.winfo_screenheight()
mov_x = screen_width / 2 - 300
mov_y = screen_height / 2 - 200
window.geometry("600x400+%d+%d" % (mov_x, mov_y))
x = window.winfo_x()
y = window.winfo_y()
window['background'] = 'white'
y2_logo = "Y2Logo.gif"
render = Tkinter.PhotoImage(file= y2_logo)
Tkinter.Label(window, image=render, borderwidth=0, highlightthickness=0, relief=Tkinter.FLAT).pack()
window.deiconify()
Internbot(window)
window.mainloop()