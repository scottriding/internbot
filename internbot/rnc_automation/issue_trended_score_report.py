from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from collections import OrderedDict
   
class IssueTrendedReport(object):

    def __init__ (self, models, number_of_rounds):
        # create final workbook
        self.models = models
        self.workbook = Workbook()
        self.total_count = 0
        self.rounds = number_of_rounds

        # font details
        self.titles_style = Font(name = 'Arial', size = 11, bold = True)
        self.general_style = Font(name = 'Arial', size = 11)
        self.second_row = Font(name = 'Arial', size = 11, bold = True, italic = True)

        # border details
        self.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        self.top_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'))
        self.middle_border = Border(left=Side(style='thick'), right=Side(style='thick'))
        self.bottom_border = Border(left=Side(style='thick'), right=Side(style='thick'), bottom=Side(style='thick'))

        # fill/highlight details
        self.grey = PatternFill("solid", fgColor="B7B7B7")
        self.key_variables = PatternFill("solid", fgColor="FFC5C5")
        self.key_models = PatternFill("solid", fgColor="D4D4D4")

        # write report
        self.build_sheets()
        self.write_model_tabs()

    def build_sheets(self):
        model_names = self.models.list_names()
        default_sheet = True
        alternate_tabs = True
        for model_name in model_names:
            if default_sheet is True:
                sheet = self.workbook.get_sheet_by_name("Sheet")
                sheet.title = str(model_name)
                default_sheet = False
            else:
                sheet = self.workbook.create_sheet(str(model_name))

            if alternate_tabs is True:
                sheet.sheet_properties.tabColor = '7F7F7F'
                alternate_tabs = False
            else:
                sheet.sheet_properties.tabColor = 'A6494A'
                alternate_tabs = True

    def write_model_tabs(self):
        model_names = self.models.list_names()
        for model_name in model_names:
            current_sheet = self.workbook.get_sheet_by_name(str(model_name))
            self.write_first_row(current_sheet)
            self.write_fields(current_sheet, model_name)

    def write_first_row(self, current_sheet):
        current_sheet["A1"].value = "Field Name"
        current_sheet["A1"].font = self.titles_style
        current_sheet["A1"].border = self.border
        current_sheet["A1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["A"].width = 35

        current_sheet["B1"].value = "Grouping"
        current_sheet["B1"].font = self.titles_style
        current_sheet["B1"].border = self.border
        current_sheet["B1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["B"].width = 39

        current_sheet["C1"].value = "Count"
        current_sheet["C1"].font = self.titles_style
        current_sheet["C1"].border = self.border
        current_sheet["C1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["C"].width = 11

        current_sheet["D1"].value = "Percent"
        current_sheet["D1"].font = self.titles_style
        current_sheet["D1"].border = self.border
        current_sheet["D1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["D"].width = 11

        current_sheet["E1"].value = "Round 1 6/13/18"
        current_sheet["E1"].font = self.titles_style
        current_sheet["E1"].border = self.border
        current_sheet["E1"].alignment = Alignment(horizontal="center", wrap_text = True)

        current_sheet.column_dimensions["E"].width = 11

    def write_fields(self, current_sheet, model_name):
        current_cell_A = 2
        current_cell_B = 2
        fields = self.models.get_fields(model_name)
        for field in fields:
            location = "A%s" % current_cell_A
            field_object = self.models.model(model_name).get(field)
            field_name = field_object.name
            current_sheet[location].value = field_name
            current_sheet[location].font = self.titles_style
            current_sheet[location].alignment = Alignment(horizontal="center")

            # merge
            group_count = field_object.length_groups()
            end = (current_cell_A + group_count) - 1
            current_sheet.merge_cells(start_column=1, end_column=1, start_row=current_cell_A, end_row = end)
            current_cell_A += group_count
            
            # write groups
            current_cell_B = self.write_groupings(current_sheet, current_cell_B, field_object)

    def write_groupings(self, current_sheet, current_cell, field_object):
        field_groups = field_object.get_groupings()
        middle_border = False
        top_border = True
        if len(field_groups) > 2:
            middle_border = True
        for field_group in field_groups:
            location_a = "A%s" % (current_cell)
            location_b = "B%s" % (current_cell)
            location_c = "C%s" % (current_cell)
            location_d = "D%s" % (current_cell)
            location_e = "E%s" % (current_cell)

            if field_group in "All Voters":
                self.total_count = field_object.get_count(field_group)
                current_sheet[location_a].fill = self.grey
                current_sheet[location_a].font = self.second_row
                current_sheet[location_b].fill = self.grey
                current_sheet[location_b].font = self.second_row
                current_sheet[location_c].fill = self.grey
                current_sheet[location_c].font = self.second_row
                current_sheet[location_d].fill = self.grey
                current_sheet[location_d].font = self.second_row
                current_sheet[location_e].fill = self.grey
                current_sheet[location_e].font = self.second_row

            current_sheet[location_b].value = field_group
            current_sheet[location_b].font = self.titles_style
            current_sheet[location_c].value = field_object.get_count(field_group)
            current_sheet[location_d].value = (float(field_object.get_count(field_group))/float(self.total_count))
            current_sheet[location_e].value = field_object.get_frequency(field_group)
            current_sheet[location_d].number_format = '0%'
            current_sheet[location_e].number_format = '0%'
            current_sheet[location_c].number_format = '#,##0'

            if top_border is True:
                current_sheet[location_a].border = self.top_border
                current_sheet[location_b].border = self.top_border
                current_sheet[location_c].border = self.top_border
                current_sheet[location_d].border = self.top_border
                current_sheet[location_e].border = self.top_border
                top_border = False
            elif middle_border is True:
                current_sheet[location_a].border = self.middle_border
                current_sheet[location_b].border = self.middle_border
                current_sheet[location_c].border = self.middle_border
                current_sheet[location_d].border = self.middle_border
                current_sheet[location_e].border = self.middle_border

            current_cell += 1

        location_a = "A%s" % (current_cell - 1)
        location_b = "B%s" % (current_cell - 1)
        location_c = "C%s" % (current_cell - 1)
        location_d = "D%s" % (current_cell - 1)
        location_e = "E%s" % (current_cell - 1)
        current_sheet[location_a].border = self.bottom_border
        current_sheet[location_b].border = self.bottom_border
        current_sheet[location_c].border = self.bottom_border
        current_sheet[location_d].border = self.bottom_border
        current_sheet[location_e].border = self.bottom_border

        return current_cell

    def insert_cols(self):
        for sheet in self.workbook.worksheets:
            sheet.insert_cols(3)

    def highlight(self, cell):
        darkest_negative = PatternFill("solid", fgColor="B80001")
        medium_dark_negative = PatternFill("solid", fgColor="CD4748")
        medium_negative = PatternFill("solid", fgColor="DF8A8C")
        medium_light_negative = PatternFill("solid", fgColor="EAB9BB")
        lightest_negative = PatternFill("solid", fgColor="F1D2D5")

        darkest_positive = PatternFill("solid", fgColor="02A747")
        medium_dark_positive = PatternFill("solid", fgColor="2AB666")
        medium_positive = PatternFill("solid", fgColor="5AC88B")
        medium_light_positive = PatternFill("solid", fgColor="B6E7CE")
        lightest_positive = PatternFill("solid", fgColor="E6F6F0")
        
        if cell.value < 0:
            if cell.value >= -0.01:
                cell.fill = lightest_negative
            elif cell.value >= -0.03:
                cell.fill = medium_light_negative
            elif cell.value >= -0.05:
                cell.fill = medium_negative
            elif cell.value >= -0.07:
                cell.fill = medium_dark_negative
            else:
                cell.fill = darkest_negative
        elif cell.value > 0:
            if cell.value <= 0.01:
                cell.fill = lightest_positive
            elif cell.value <= 0.03:
                cell.fill = medium_light_positive
            elif cell.value <= 0.05:
                cell.fill = medium_positive
            elif cell.value <= 0.07:
                cell.fill = medium_dark_positive
            else:
                cell.fill = darkest_positive

    def save(self, path_to_output):
        self.workbook.save(path_to_output)
        