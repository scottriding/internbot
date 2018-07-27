from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from collections import OrderedDict
   
class TrendedScoreReport(object):

    def __init__ (self, all_workbook_details, path_to_output, number_of_rounds):
        self.total_count = 0
        self.rounds = int(number_of_rounds)
        self.date = all_workbook_details
        
        # font details
        self.titles_style = Font(name='Calibri (Body)', size=11, bold=True)
        self.general_style = Font(name='Calibri (Body)', size=11)
        self.second_row = Font(name='Calibri (Body)', size=11, bold=True, italic=True)

        # border details
        self.border_all = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        self.top_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'))
        self.middle_border = Border(left=Side(style='thick'), right=Side(style='thick'))
        self.bottom_border = Border(left=Side(style='thick'), right=Side(style='thick'),bottom=Side(style='thick'))

        # fill/highlight details
        self.grey = PatternFill("solid", fgColor="B7B7B7")

        # calculate the excel alphabet from G to ZZ
        alphabet = []
        for letter in range(71, 91):
            alphabet.append(chr(letter))

        self.extend_alphabet = []
        self.extend_alphabet.extend(alphabet)
        complete_alphabet = []
        for letter in range(65, 91):
            complete_alphabet.append(chr(letter))
        index = 0
        while index < len(complete_alphabet):
            for letter in complete_alphabet:
                double_letters = "%s%s" % (complete_alphabet[index], letter)
                self.extend_alphabet.append(double_letters)
            index += 1

        # build reports
        self.build_workbooks(all_workbook_details, path_to_output)

    def build_workbooks(self, all_workbook_details, path_to_output):
        workbook_names = all_workbook_details.list_workbook_names()
        for workbook_name in workbook_names:
            new_workbook = Workbook()
            self.build_sheets(new_workbook, all_workbook_details.get_workbook(workbook_name))
            save_path = path_to_output + "/" + workbook_name + ".xlsx"
            new_workbook.save(save_path)

    def build_sheets(self, workbook, workbook_details):
        default_sheet = False
        if self.workbook.get_sheet_by_name("Sheet") is not None:
            default_sheet = True

        sheet_names = workbook_details.list_sheet_names()
        for sheet_name in sheet_names:
            if default_sheet is True:
                sheet = workbook.get_sheet_by_name("Sheet")
                sheet.title = sheet_name
                default_sheet = False
            else:
                sheet = workbook.create_sheet(sheet_name)
            self.write_sheet(sheet, workbook_details.get_sheet(sheet_name))

    def write_sheet(self, sheet, sheet_details):
        self.write_first_row(sheet)
        current_row_field = 2
        current_row_grouping = 2
        field_names = sheet_details.list_field_names()
        for field_name in field_names:
            location = "A%s" % current_row_field
            sheet[location].value = field_name
            sheet[location].font = self.titles_style
            sheet[location].alignment = Alignment(horizontal="center")
            field_object = sheet_details.get_field(field_name)
            group_count = field_object.grouping_count()
            end = (current_row_field + group_count) - 1
            sheet.merge_cells(start_column=1, end_column=1, start_row=current_row_field, end_row=end)
            current_row_field += group_count
            current_row_grouping = self.write_groupings(sheet, current_row_grouping, field_object)
        self.write_difference(sheet, sheet_details)
            

    def write_first_row(self, current_sheet):
        current_sheet["A1"].value = "Field Name"
        current_sheet["A1"].font = self.titles_style
        current_sheet["A1"].border = self.border_all
        current_sheet["A1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["A"].width = 33

        current_sheet["B1"].value = "Grouping"
        current_sheet["B1"].font = self.titles_style
        current_sheet["B1"].border = self.border_all
        current_sheet["B1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["B"].width = 29

        current_sheet["C1"].value = "Count"
        current_sheet["C1"].font = self.titles_style
        current_sheet["C1"].border = self.border_all
        current_sheet["C1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["C"].width = 11

        current_sheet["D1"].value = "Percent"
        current_sheet["D1"].font = self.titles_style
        current_sheet["D1"].border = self.border_all
        current_sheet["D1"].alignment = Alignment(horizontal="center")

        current_sheet.column_dimensions["D"].width = 11

        current_sheet["E1"].value = "Current Round-Previous Round"
        current_sheet["E1"].font = self.titles_style
        current_sheet["E1"].border = self.border_all
        current_sheet["E1"].alignment = Alignment(horizontal="center", wrap_text = True)

        current_sheet.column_dimensions["E"].width = 13

        current_sheet["F1"].value = "Current Round-First Round"
        current_sheet["F1"].font = self.titles_style
        current_sheet["F1"].border = self.border_all
        current_sheet["F1"].alignment = Alignment(horizontal="center", wrap_text = True)

        current_sheet.column_dimensions["F"].width = 13

        round_iteration = self.rounds
        index = 0
        while round_iteration > 0:
            header = "Round %s %s" % (round_iteration, self.date.round_date(round_iteration))
            column = self.extend_alphabet[index]
            cell = "%s1" % column
            current_sheet[cell].value = header
            current_sheet[cell].border = self.border_all
            current_sheet[cell].font = self.titles_style
            current_sheet[cell].alignment = Alignment(horizontal="center", wrap_text = True)
            round_iteration -= 1
            index += 1

    def write_groupings(self, current_sheet, current_row, field_object):
        grouping_names = field_object.list_grouping_names()
        middle_border = False
        top_border = True
        if len(grouping_names) > 2:
            middle_border = True
        for grouping_name in grouping_names:
            grouping = field_object.get_grouping(grouping_name)

            field_cell = "A%s" % (current_row)
            grouping_cell = "B%s" % (current_row)
            count_cell = "C%s" % (current_row)
            percent_cell = "D%s" % (current_row)
            prev_diff_cell = "E%s" % (current_row)
            first_diff_cell = "F%s" % (current_row)

            if current_row == 2:
                self.total_count = grouping.count
                current_sheet[field_cell].fill = self.grey
                current_sheet[field_cell].font = self.second_row
                current_sheet[grouping_cell].fill = self.grey
                current_sheet[grouping_cell].font = self.second_row
                current_sheet[count_cell].fill = self.grey
                current_sheet[count_cell].font = self.second_row
                current_sheet[percent_cell].fill = self.grey
                current_sheet[percent_cell].font = self.second_row
                current_sheet[prev_diff_cell].font = self.second_row
                current_sheet[first_diff_cell].font = self.second_row

            ## grouping name
            current_sheet[grouping_cell].value = grouping_name
            current_sheet[grouping_cell].font = self.titles_style
            
            # grouping count
            current_sheet[count_cell].value = grouping.count
            current_sheet[count_cell].number_format = '#,##0'
            
            # grouping percent
            percent = grouping.count/self.total_count
            current_sheet[percent_cell].value = percent
            current_sheet[percent_cell].number_format = '0%'
            
            # difference columns
            current_sheet[prev_diff_cell].number_format = '0%'
            current_sheet[first_diff_cell].number_format = '0%'

            # round frequencies
            index = 0
            round_iteration = self.rounds
            while round_iteration > 0:
                round_cell = "%s%s" % (self.extend_alphabet[index], current_row)
                current_sheet[round_cell].value = grouping.round_frequency(round_iteration)
                current_sheet[round_cell].number_format = '0%'
                if current_row == 2:
                    current_sheet[round_cell].font = self.second_row
                    current_sheet[round_cell].fill = self.grey
                if top_border is True:
                    current_sheet[round_cell].border = self.top_border
                elif middle_border is True:
                    current_sheet[round_cell].border = self.middle_border
                round_iteration -= 1
                index += 1

            # borders
            if top_border is True:
                current_sheet[field_cell].border = self.top_border
                current_sheet[grouping_cell].border = self.top_border
                current_sheet[count_cell].border = self.top_border
                current_sheet[percent_cell].border = self.top_border
                current_sheet[prev_diff_cell].border = self.top_border
                current_sheet[first_diff_cell].border = self.top_border
                top_border = False
            elif middle_border is True:
                current_sheet[field_cell].border = self.middle_border
                current_sheet[grouping_cell].border = self.middle_border
                current_sheet[count_cell].border = self.middle_border
                current_sheet[percent_cell].border = self.middle_border
                current_sheet[prev_diff_cell].border = self.middle_border
                current_sheet[first_diff_cell].border = self.middle_border

            current_row += 1

        # bottom borders
        field_cell = "A%s" % (current_row - 1)
        grouping_cell = "B%s" % (current_row - 1)
        count_cell = "C%s" % (current_row - 1)
        percent_cell = "D%s" % (current_row - 1)
        prev_diff_cell = "E%s" % (current_row - 1)
        first_diff_cell = "F%s" % (current_row - 1)
        current_sheet[field_cell].border = self.bottom_border
        current_sheet[grouping_cell].border = self.bottom_border
        current_sheet[count_cell].border = self.bottom_border
        current_sheet[percent_cell].border = self.bottom_border
        current_sheet[prev_diff_cell].border = self.bottom_border
        current_sheet[first_diff_cell].border = self.bottom_border

        # bottom round borders
        index = 0
        round_iteration = self.rounds
        while round_iteration > 0:
            round_cell = "%s%s" % (self.extend_alphabet[index], (current_row - 1))
            current_sheet[round_cell].border = self.bottom_border
            round_iteration -= 1
            index += 1

        return current_row

    def write_difference(self, sheet, current):
        current_row = 2
        field_names = current.list_field_names()
        for field_name in field_names:
            field = current.get_field(field_name)
            grouping_names = field.list_grouping_names()
            for grouping_name in grouping_names:
                prev_diff_cell = "E%s" % (current_row)
                first_diff_cell = "F%s" % (current_row)

                current_round_cell = "G%s" % (current_row)
                prev_round_cell = "H%s" % (current_row)
                first_round_cell = "%s%s" % (self.extend_alphabet[self.rounds - 1], current_row)

                if self.rounds == 1:
                    sheet[prev_diff_cell].value = "--%"
                    sheet[first_diff_cell].value = "--%"
                else:
                    sheet[prev_diff_cell].value = "=%s - %s" % (current_round_cell, prev_round_cell)
                    sheet[prev_diff_cell].fill = self.highlight(sheet[current_round_cell].value, sheet[prev_round_cell].value)
                    sheet[first_diff_cell].value = "=%s - %s" % (current_round_cell, first_round_cell)
                    sheet[first_diff_cell].fill = self.highlight(sheet[current_round_cell].value, sheet[first_round_cell].value)
                current_row += 1

    def highlight(self, first_value, second_value):
        darkest_negative = PatternFill("solid", fgColor="B80001")
        medium_dark_negative = PatternFill("solid", fgColor="CD4748")
        medium_negative = PatternFill("solid", fgColor="DF8A8C")
        medium_light_negative = PatternFill("solid", fgColor="EAB9BB")
        lightest_negative = PatternFill("solid", fgColor="F1D2D5")
 
        darkest_positive = PatternFill("solid", fgColor="02A747")
        medium_dark_positive = PatternFill("solid", fgColor="2AB666")
        medium_positive = PatternFill("solid", fgColor="5AC88B")
        medium_light_positive = PatternFill("solid", fgColor="B6E7CE")
        lightest_positive = PatternFill("solid", fgColor="E6F6F0")
        
        grey = PatternFill("solid",fgColor="E6E6E6")
        if first_value is None or second_value is None:
            return grey
        
        result = first_value - second_value
        if result < 0:
            if result >= -0.01:
                return lightest_negative
            elif result >= -0.03:
                return medium_light_negative
            elif result >= -0.05:
                return medium_negative
            elif result >= -0.07:
                return medium_dark_negative
            elif result < -0.07:
                return darkest_negative
            else:
                return grey
        elif result > 0:
            if result <= 0.01:
                return lightest_positive
            elif result <= 0.03:
                return medium_light_positive
            elif result <= 0.05:
                return medium_positive
            elif result <= 0.07:
                return medium_dark_positive
            elif result > 0.07:
                return darkest_positive
            else:
                return grey
        else:
            return grey
        