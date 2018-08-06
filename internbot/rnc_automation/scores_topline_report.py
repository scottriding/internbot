from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from collections import OrderedDict
   
class ScoresToplineReport(object):

    def __init__ (self, models, report_location, number_of_rounds):
        # create final workbook
        self.models = models
        self.report_location = report_location
        self.rounds = int(number_of_rounds)
        self.workbook = Workbook()

        # font details
        self.titles_style = Font(name = 'Calibri (Body)', size = 11, bold = True)
        self.general_style = Font(name = 'Calibri (Body)', size = 11)
        self.net_score_style = Font(name = 'Calibri (Body)', size = 11, italic = True)
        self.key_title_style = Font(name = "Calibri (Body)", size = 8, bold = True)
        self.key_style = Font(name = "Calibri (Body)", size = 8)
        self.net_style = Font(name = 'Calibri (Body)', size = 8, italic = True)
        
        # border details
        self.all_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        self.top_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'))
        self.middle_border = Border(left=Side(style='thick'), right=Side(style='thick'))
        self.bottom_border = Border(left=Side(style='thick'), right=Side(style='thick'), bottom=Side(style='thick'))

        # fill/highlight details
        self.grey = PatternFill("solid", fgColor="B7B7B7")
        self.key_variables = PatternFill("solid", fgColor="FFC5C5")
        self.key_models = PatternFill("solid", fgColor="D4D4D4")

        # calculate the excel alphabet from D to ZZ
        alphabet = []
        for letter in range(68, 91):
            alphabet.append(chr(letter))

        self.extend_alphabet = []
        self.extend_alphabet.extend(alphabet)
        complete_alphabet = []
        for letter in range(65, 91):
            complete_alphabet.append(chr(letter))
        index = 0
        while index < len(complete_alphabet):
            for letter in complete_alphabet:
                double_letters = "%s%s" % (complete_alphabet[index], letter)
                self.extend_alphabet.append(double_letters)
            index += 1

        # write report
        self.build_sheets()

    def build_sheets(self):
        new_title = "Key %s" % (self.report_location)
        if self.workbook.get_sheet_by_name("Sheet") is not None:
            sheet = self.workbook.get_sheet_by_name('Sheet')
            sheet.title = new_title
        else:
            sheet = self.workbook.create_sheet(str(new_title))

        new_sheet = self.report_location
        self.workbook.create_sheet(str(new_sheet))
        self.write_key(sheet)
        self.write_report(self.workbook.get_sheet_by_name(self.report_location))

    def write_key(self, key_sheet):
        self.write_key_titles(key_sheet)
        self.write_key_models(key_sheet)

    def write_key_titles(self, key_sheet):
        current_title_cell = "B1"
        current_model_cell = "A2"
        current_refer_cell = "B2"
        previous_title_cell = "E1"
        previous_model_cell = "D2"
        previous_refer_cell = "E2"

        key_sheet[current_title_cell].value = "Round %s" % str(self.rounds)
        key_sheet[current_title_cell].font = self.titles_style
        key_sheet[current_title_cell].border = self.all_border
        key_sheet[current_model_cell].value = "MODEL"
        key_sheet[current_model_cell].font = self.key_title_style
        key_sheet[current_model_cell].border = self.all_border
        key_sheet[current_refer_cell].value = "SURVEY QUESTION REFERENCE"
        key_sheet[current_refer_cell].font = self.key_title_style
        key_sheet[current_refer_cell].border = self.all_border

        key_sheet.column_dimensions["A"].width = 26
        key_sheet.column_dimensions["B"].width = 220

        if self.rounds > 1:
            key_sheet[previous_title_cell].value = "Round %s" % str(self.rounds - 1)
            key_sheet[previous_title_cell].font = self.titles_style
            key_sheet[previous_title_cell].border = self.all_border
            key_sheet[previous_model_cell].value = "MODEL"
            key_sheet[previous_model_cell].font = self.key_title_style
            key_sheet[previous_model_cell].border = self.all_border
            key_sheet[previous_refer_cell].value = "SURVEY QUESTION REFERENCE"
            key_sheet[previous_refer_cell].font = self.key_title_style
            key_sheet[previous_refer_cell].border = self.all_border

            key_sheet.column_dimensions["D"].width = 26
            key_sheet.column_dimensions["E"].width = 220

    def write_key_models(self, key_sheet):
        current_round_row = 3
        previous_round_row = 3
        model_names = self.models.list_model_names()
        for model_name in model_names:
            current_model_cell = "A%s" % current_round_row
            current_refer_cell = "B%s" % current_round_row
            previous_model_cell = "D%s" % previous_round_row
            previous_refer_cell = "E%s" % previous_round_row
        
            current = self.models.get_model(model_name)
            
            # net model name
            key_sheet[current_model_cell].value = current.name
            key_sheet[current_model_cell].font = self.net_style
            key_sheet[current_model_cell].border = self.all_border
            key_sheet[current_model_cell].fill = self.key_models
        
            # net model survey question reference
            key_sheet[current_refer_cell].font = self.net_style
            key_sheet[current_refer_cell].border = self.all_border
            key_sheet[current_refer_cell].fill = self.key_models

            current_round_row = self.write_current_key_variables(key_sheet, current, current_round_row)

            if self.rounds > 1:
                # check if model is NA or not in this particular round
                variable_names = current.list_variable_names()
                freq_to_check = current.get_variable(variable_names[0]).round_weighted_freq(self.rounds-1)
                if freq_to_check != "NA" and freq_to_check != "":
                    # net model name
                    key_sheet[previous_model_cell].value = current.name
                    key_sheet[previous_model_cell].font = self.net_style
                    key_sheet[previous_model_cell].border = self.all_border
                    key_sheet[previous_model_cell].fill = self.key_models
    
                    # net model survey question reference
                    key_sheet[previous_refer_cell].font = self.net_style
                    key_sheet[previous_refer_cell].border = self.all_border
                    key_sheet[previous_refer_cell].fill = self.key_models
                    previous_round_row = self.write_previous_key_variables(key_sheet, current, previous_round_row)

    def write_current_key_variables(self, key_sheet, model, current_row):
        current_row += 1
        start = current_row
        variable_names = model.list_variable_names()
        middle_border = False
        first_border = True
        if len(variable_names) > 2:
            middle_border = True
        for variable_name in variable_names:
            current_model_cell = "A%s" % current_row
            
            # variable name
            key_sheet[current_model_cell].value = variable_name
            key_sheet[current_model_cell].font = self.key_style
            key_sheet[current_model_cell].fill = self.key_variables
            if first_border == True:
                key_sheet[current_model_cell].border = self.top_border
            elif middle_border == True:
                key_sheet[current_model_cell].border = self.middle_border

            first_border = False
            current_row += 1
        
        key_sheet[current_model_cell].border = self.bottom_border

        end = current_row - 1
        length = len(variable_names)
        self.write_current_key_reference(key_sheet, model, start, end, length)

        return current_row

    def write_previous_key_variables(self, key_sheet, model, current_row):
        current_row += 1
        start = current_row
        variable_names = model.list_variable_names()
        middle_border = False
        first_border = True
        if len(variable_names) > 2:
            middle_border = True
        for variable_name in variable_names:
            previous_model_cell = "D%s" % current_row

            # variable name
            key_sheet[previous_model_cell].value = variable_name
            key_sheet[previous_model_cell].font = self.key_style
            key_sheet[previous_model_cell].fill = self.key_variables
            if first_border == True:
                key_sheet[previous_model_cell].border = self.top_border
            elif middle_border == True:
                key_sheet[previous_model_cell].border = self.middle_border

            first_border = False
            current_row += 1
        
        key_sheet[previous_model_cell].border = self.bottom_border
        
        end = current_row - 1
        length = len(variable_names)
        self.write_previous_key_reference(key_sheet, model, start, end, length)

        return current_row
        
    def write_current_key_reference(self, key_sheet, model, start, end, length):
        # model survey question reference
        current_refer_cell = "B%s" % start
        key_sheet.merge_cells(start_column=2, end_column=2, start_row=start, end_row = end)
        key_sheet[current_refer_cell].value = model.description
        key_sheet[current_refer_cell].font = self.key_style
        key_sheet[current_refer_cell].fill = self.key_variables

        first_border = True
        middle_border = False
        if length > 2:
            middle_border = True
        while start <= end:
            current_refer_cell = "B%s" % start
            if first_border == True:
                key_sheet[current_refer_cell].border = self.top_border
                first_border = False
            elif middle_border == True:
                key_sheet[current_refer_cell].border = self.middle_border
                middle_border = False
            start += 1
        key_sheet[current_refer_cell].border = self.bottom_border

    def write_previous_key_reference(self, key_sheet, model, start, end, length):
        # model survey question reference
        previous_refer_cell = "E%s" % start
        key_sheet.merge_cells(start_column=5, end_column=5, start_row=start, end_row = end)
        key_sheet[previous_refer_cell].value = model.description
        key_sheet[previous_refer_cell].font = self.key_style
        key_sheet[previous_refer_cell].fill = self.key_variables

        first_border = True
        middle_border = False
        if length > 2:
            middle_border = True
        while start <= end:
            previous_refer_cell = "E%s" % start
            if first_border == True:
                key_sheet[previous_refer_cell].border = self.top_border
                first_border = False
            elif middle_border == True:
                key_sheet[previous_refer_cell].border = self.middle_border
                middle_border = False
            start += 1
        key_sheet[previous_refer_cell].border = self.bottom_border

    def write_report(self, scores_sheet):
        self.write_report_titles(scores_sheet)
        self.write_report_models(scores_sheet)
        self.write_net_differences(scores_sheet)

    def write_report_titles(self, score_sheet):
        ## hard coding the header columns
        current_prev_location = "B1"
        current_first_location = "C1"
        model_header_location = "A2"

        score_sheet[current_prev_location].value = "Current Round - Previous Round"
        score_sheet[current_prev_location].font = self.titles_style
        score_sheet[current_prev_location].border = self.all_border
        score_sheet[current_prev_location].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
        score_sheet["B2"].border = self.all_border

        score_sheet.column_dimensions["B"].width = 14

        score_sheet[current_first_location].value = "Current Round - First Round"
        score_sheet[current_first_location].font = self.titles_style
        score_sheet[current_first_location].border = self.all_border
        score_sheet[current_first_location].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
        score_sheet["C2"].border = self.all_border

        score_sheet.column_dimensions["C"].width = 14

        score_sheet[model_header_location].value = "Model"
        score_sheet[model_header_location].font = self.titles_style
        score_sheet[model_header_location].border = self.all_border

        score_sheet.column_dimensions["A"].width = 45

        # Each column in the second row get automated headers based on number of rounds
        iteration = self.rounds
        index = 0
        start = 4
        end = 4
        while iteration > 0:
            weighted_cell = "%s2" % (self.extend_alphabet[index])
            unweighted_cell = "%s2" % (self.extend_alphabet[index + 1])
            score_sheet[weighted_cell].value = "Round %s TOW" % iteration
            score_sheet[weighted_cell].font = self.titles_style
            score_sheet[weighted_cell].border = self.all_border
            score_sheet[unweighted_cell].value = "Round %s %s" % (iteration, self.models.round_date(iteration))
            score_sheet[unweighted_cell].font = self.titles_style
            score_sheet[unweighted_cell].border = self.all_border
            score_sheet.column_dimensions[self.extend_alphabet[index]].width = 14
            score_sheet.column_dimensions[self.extend_alphabet[index + 1]].width = 14
            end += 2
            iteration -= 1
            index += 2

        # merge for state name
        score_sheet["D1"].value = self.report_location.upper()
        score_sheet["D1"].font = Font(name = 'Calibri (Body)', size = 12, bold = True)
        score_sheet["D1"].alignment = Alignment(horizontal='center', vertical='center')
        score_sheet.merge_cells(start_column=start, end_column=end - 1, start_row=1, end_row = 1)

    def write_report_models(self, score_sheet):
        current_row = 3
        # write net level
        model_names = self.models.list_model_names()
        for model_name in model_names:
            if model_name == "Turnout General":
                description_cell = "A%s" % current_row
                score_sheet[description_cell].value = model_name
                score_sheet[description_cell].font = self.net_score_style
                score_sheet[description_cell].border = self.all_border

                unweighted_frequencies = self.models.get_model(model_name).get_variable(model_name).unweighted_frequencies()

                iteration = self.rounds
                index = 0
                while iteration > 0:
                    weighted_cell = "%s%s" % (self.extend_alphabet[index], current_row)
                    unweighted_cell = "%s%s" % (self.extend_alphabet[index + 1], current_row)
                    score_sheet[weighted_cell].fill = self.grey
                    score_sheet[weighted_cell].border = self.all_border
                    score_sheet[unweighted_cell].value = float(unweighted_frequencies[iteration - 1])
                    score_sheet[unweighted_cell].number_format = '0%'
                    score_sheet[unweighted_cell].border = self.all_border
                    iteration -= 1
                    index += 2

                current_row += 1
            else:
                description_cell = "A%s" % current_row
                current = self.models.get_model(model_name)
                score_sheet[description_cell].value = model_name
                score_sheet[description_cell].font = self.net_score_style
                score_sheet[description_cell].border = self.all_border
                iteration = self.rounds
                index = 0
                while iteration > 0:
                    net_weighted_cell = "%s%s" % (self.extend_alphabet[index], current_row)
                    net_unweighted_cell = "%s%s" % (self.extend_alphabet[index + 1], current_row)
                    last_row = current_row + 2
                    first_row = current_row + 1
                    if self.is_na(current, iteration) == True:
                        score_sheet[net_weighted_cell].fill = self.grey
                        score_sheet[net_unweighted_cell].fill = self.grey
                    else:
                        weight_difference_1 = "%s%s" % (self.extend_alphabet[index], first_row)
                        weight_difference_2 = "%s%s" % (self.extend_alphabet[index], last_row)
                        unweight_difference_1 = "%s%s" % (self.extend_alphabet[index + 1], first_row)
                        unweight_difference_2 = "%s%s" % (self.extend_alphabet[index + 1], last_row)
                        score_sheet[net_weighted_cell].value = "=%s - %s" % (weight_difference_1, weight_difference_2)
                        score_sheet[net_weighted_cell].number_format = '0%'
                        score_sheet[net_unweighted_cell].value = "=%s - %s" % (unweight_difference_1, unweight_difference_2)
                        score_sheet[net_unweighted_cell].number_format = '0%'
                        end_column = self.extend_alphabet[index + 1]
                    
                    score_sheet[net_weighted_cell].font = self.net_score_style
                    score_sheet[net_weighted_cell].border = self.all_border
                    score_sheet[net_unweighted_cell].font = self.net_score_style
                    score_sheet[net_unweighted_cell].border = self.all_border
                    iteration -= 1
                    index += 2

                # write variable level
                current_row = self.write_report_variable(score_sheet, current, current_row)

    def write_report_variable(self, score_sheet, model, current_row):
        variable_names = model.list_variable_names()
        current_row += 1
        top_border = True
        middle_border = False
        if len(variable_names) > 2:
            middle_border = True
        for variable_name in variable_names:
            name_cell = "A%s" % current_row
            score_sheet[name_cell].value = variable_name
            if top_border == True:
                score_sheet[name_cell].border = self.top_border
            elif middle_border == True:
                score_sheet[name_cell].border = self.middle_border
            variable = model.get_variable(variable_name)
            iteration = self.rounds
            index = 0
            while iteration > 0:
                weighted_cell = "%s%s" % (self.extend_alphabet[index], current_row)
                unweighted_cell = "%s%s" % (self.extend_alphabet[index + 1], current_row)
                frequency_weighted = variable.round_weighted_freq(iteration)
                if frequency_weighted == "NA" or frequency_weighted == "":
                    score_sheet[weighted_cell].fill = self.grey
                else:
                    score_sheet[weighted_cell].value = float(frequency_weighted)
                    score_sheet[weighted_cell].font = self.general_style
                    score_sheet[weighted_cell].number_format = '0%'
                frequency_unweighted = variable.round_unweighted_freq(iteration)
                if frequency_unweighted == "NA" or frequency_unweighted == "":
                    score_sheet[unweighted_cell].fill = self.grey
                else:
                    score_sheet[unweighted_cell].value = float(frequency_unweighted)
                    score_sheet[unweighted_cell].font = self.general_style
                    score_sheet[unweighted_cell].number_format = '0%'
                if top_border == True:
                    score_sheet[weighted_cell].border = self.top_border
                    score_sheet[unweighted_cell].border = self.top_border
                elif middle_border == True:
                    score_sheet[weighted_cell].border = self.middle_border
                    score_sheet[unweighted_cell].border = self.middle_border
                iteration -= 1
                index += 2
            top_border = False
            current_row += 1
        bottom_cell = current_row - 1
        index = 0
        iteration = self.rounds
        while iteration > 0:
            weighted_cell = "%s%s" % (self.extend_alphabet[index], bottom_cell)
            unweighted_cell = "%s%s" % (self.extend_alphabet[index + 1], bottom_cell)
            score_sheet[weighted_cell].border = self.bottom_border
            score_sheet[unweighted_cell].border = self.bottom_border
            iteration -= 1
            index += 2
        name_cell = "A%s" % str(bottom_cell)
        score_sheet[name_cell].border = self.bottom_border

        return current_row

    def write_net_differences(self, score_sheet):
        current_row = 3

        model_names = self.models.list_model_names()
        for model_name in model_names:
            prev_diff_cell = "B%s" % current_row
            first_diff_cell = "C%s" % current_row
            if model_name == "Turnout General":
                score_sheet[prev_diff_cell].fill = self.grey
                score_sheet[first_diff_cell].fill = self.grey
                score_sheet[prev_diff_cell].border = self.all_border
                score_sheet[first_diff_cell].border = self.all_border
                current_row += 1
            elif self.rounds == 1:
                score_sheet[prev_diff_cell].border = self.all_border
                score_sheet[first_diff_cell].border = self.all_border
                score_sheet[prev_diff_cell].value = "--%"
                score_sheet[first_diff_cell].value = "--%"
                current = self.models.get_model(model_name)
                current_row = self.write_variable_differences(score_sheet, current, current_row)
            else:
                # more than 1 round and not a Turnout General
                current_round_cell = "E%s" % current_row
                previous_round_cell = "G%s" % current_row

                current = self.models.get_model(model_name)
                first_round_cell = self.find_first_round(score_sheet, current, current_row)

                # check if first round is nonexistent
                no_first = False
                if first_round_cell == previous_round_cell:
                    no_first = True

                # check if previous round is empty
                no_prev = False
                if score_sheet[previous_round_cell].value is None:
                    no_prev = True

                # check if current round is empty
                no_current = False
                if score_sheet[current_round_cell].value is None:
                    no_current = True

                ### figure out edge cases here ###
                ## a netmodel with no models
                if len(current.list_variable_names()) == 1:
                    pass
                ## a model with no current frequency -- a model taken off current round
                elif no_current == True:
                    score_sheet[prev_diff_cell].border = self.all_border
                    score_sheet[first_diff_cell].border = self.all_border
                    score_sheet[prev_diff_cell].value = "--%"
                    score_sheet[first_diff_cell].value = "--%"
                ## a model only added to the current round ##
                elif no_prev == True and no_first == True:
                    score_sheet[prev_diff_cell].border = self.all_border
                    score_sheet[first_diff_cell].border = self.all_border
                    score_sheet[prev_diff_cell].value = "--%"
                    score_sheet[first_diff_cell].value = "--%"
                ## a model in first and current round only
                elif no_prev == True and no_first == False:
                    score_sheet[prev_diff_cell].border = self.all_border
                    score_sheet[prev_diff_cell].value = "--%"
                    self.net_differences(score_sheet, current, current_row, first_diff_cell, current_round_cell, first_round_cell)
                ## a model in previous and current round only
                elif no_prev == False and no_first == True:
                    self.net_differences(score_sheet, current, current_row, first_diff_cell, current_round_cell, first_round_cell)
                    self.net_differences(score_sheet, current, current_row, prev_diff_cell, current_round_cell, previous_round_cell)
                ## no current edge cases apply ##
                else:
                    self.net_differences(score_sheet, current, current_row, first_diff_cell, current_round_cell, first_round_cell)
                    self.net_differences(score_sheet, current, current_row, prev_diff_cell, current_round_cell, previous_round_cell)

                ### calculate variable differences ###
                current_row = self.write_variable_differences(score_sheet, current, current_row)

    def find_first_round(self, score_sheet, model, current_row):
        index = (self.rounds * 2) - 1
        first_cell = "G%s" % current_row
        while index > 3:
            first_column = self.extend_alphabet[index]
            test_cell = "%s%s" % (first_column, current_row)
            if score_sheet[test_cell].value == None:
                index -= 2
            else:
                return test_cell
        return first_cell
    
    def net_differences(self, score_sheet, current, current_row, results_cell, current_round_cell, calculate_round_cell):
        # calculate current round formula solution
        formula_current = score_sheet[current_round_cell].value
        formula_current = formula_current.replace("=", "")
        cells_current = formula_current.split(" - ")
        cell_one = cells_current[0]
        cell_two = cells_current[1]
        cell_one_value = score_sheet[cell_one].value
        cell_two_value = score_sheet[cell_two].value
        formula_current_solution = cell_one_value - cell_two_value

        # calculate first/prev round formula solution
        formula_first = score_sheet[calculate_round_cell].value
        formula_first = formula_first.replace("=", "")
        cells_first = formula_first.split(" - ")
        cell_one = cells_first[0]
        cell_two = cells_first[1]
        cell_one_value = score_sheet[cell_one].value
        cell_two_value = score_sheet[cell_two].value
        formula_first_soluation = cell_one_value - cell_two_value

        # formatting
        score_sheet[results_cell].value = "=%s - %s" % (current_round_cell, calculate_round_cell)
        score_sheet[results_cell].fill = self.highlight(formula_current_solution, formula_first_soluation)
        score_sheet[results_cell].border = self.all_border
        score_sheet[results_cell].number_format = '0%'

    def write_variable_differences(self, score_sheet, model, current_row):
        current_row += 1
        variable_names = model.list_variable_names()
        top_border = True
        middle_border = False
        if len(variable_names) > 2:
            middle_border = True
        for variable_name in variable_names:
            prev_diff_cell = "B%s" % current_row
            first_diff_cell = "C%s" % current_row

            ## borders ##
            if top_border == True:
                score_sheet[prev_diff_cell].border = self.top_border
                score_sheet[first_diff_cell].border = self.top_border
                top_border = False
            elif middle_border == True:
                score_sheet[prev_diff_cell].border = self.middle_border
                score_sheet[first_diff_cell].border = self.middle_border

            if self.rounds == 1:
                score_sheet[prev_diff_cell].value = "--%"
                score_sheet[first_diff_cell].value = "--%"
            else:
                # more than 1 round and not a Turnout General
                current_round_cell = "E%s" % current_row
                previous_round_cell = "G%s" % current_row

                first_round_cell = self.find_first_round(score_sheet, model, current_row)
                
                # check if first round is nonexistent
                no_first = False
                if first_round_cell == previous_round_cell:
                    no_first = True

                # check if previous round is empty
                no_prev = False
                if score_sheet[previous_round_cell].value is None:
                    no_prev = True

                # check if current round is empty
                no_current = False
                if score_sheet[current_round_cell].value is None:
                    no_current = True

                ### figure out edge cases here ###
                ## a variable with no current frequency -- a model taken off current round
                if no_current == True:
                    score_sheet[prev_diff_cell].value = "--%"
                    score_sheet[first_diff_cell].value = "--%"
                ## a variable only added to the current round ##
                elif no_prev == True and no_first == True:
                    score_sheet[prev_diff_cell].value = "--%"
                    score_sheet[first_diff_cell].value = "--%"
                ## a variable in first and current round only
                elif no_prev == True and no_first == False:
                    score_sheet[prev_diff_cell].value = "--%"
                    score_sheet[first_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[first_round_cell].value)
                    score_sheet[first_diff_cell].value = "=%s - %s" % (current_round_cell, first_round_cell)
                    score_sheet[first_diff_cell].number_format = '0%'
                ## a variable in previous and current round only
                elif no_prev == False and no_first == True:
                    score_sheet[first_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[first_round_cell].value)
                    score_sheet[first_diff_cell].value = "=%s - %s" % (current_round_cell, first_round_cell)
                    score_sheet[first_diff_cell].number_format = '0%'
                    score_sheet[prev_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[previous_round_cell].value)
                    score_sheet[prev_diff_cell].value = "=%s - %s" % (current_round_cell, previous_round_cell)
                    score_sheet[prev_diff_cell].number_format = '0%'
                ## no current edge cases apply ##
                else:
                    score_sheet[prev_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[previous_round_cell].value)
                    score_sheet[prev_diff_cell].value = "=%s - %s" % (current_round_cell, previous_round_cell)
                    score_sheet[prev_diff_cell].number_format = '0%'
                    score_sheet[first_diff_cell].fill = self.highlight(score_sheet[current_round_cell].value, score_sheet[first_round_cell].value)
                    score_sheet[first_diff_cell].value = "=%s - %s" % (current_round_cell, first_round_cell)
                    score_sheet[first_diff_cell].number_format = '0%'
            current_row += 1
        score_sheet[prev_diff_cell].border = self.bottom_border
        score_sheet[first_diff_cell].border = self.bottom_border
        return current_row
    
    def is_na(self, model, current_round):
        variables = model.list_variable_names()
        unweighted_freq = model.get_variable(variables[0]).round_unweighted_freq(current_round)
        if unweighted_freq == "NA" or unweighted_freq == "":
            return True
        else:
            return False

    def highlight(self, first_value, second_value):
        darkest_negative = PatternFill("solid", fgColor="B80001")
        medium_dark_negative = PatternFill("solid", fgColor="CD4748")
        medium_negative = PatternFill("solid", fgColor="DF8A8C")
        medium_light_negative = PatternFill("solid", fgColor="EAB9BB")
        lightest_negative = PatternFill("solid", fgColor="F1D2D5")
 
        darkest_positive = PatternFill("solid", fgColor="02A747")
        medium_dark_positive = PatternFill("solid", fgColor="2AB666")
        medium_positive = PatternFill("solid", fgColor="5AC88B")
        medium_light_positive = PatternFill("solid", fgColor="B6E7CE")
        lightest_positive = PatternFill("solid", fgColor="E6F6F0")
        
        grey = PatternFill("solid",fgColor="E6E6E6")
        if first_value is None or second_value is None:
            return grey
        
        result = first_value - second_value
        if result < 0:
            if result >= -0.01:
                return lightest_negative
            elif result >= -0.03:
                return medium_light_negative
            elif result >= -0.05:
                return medium_negative
            elif result >= -0.07:
                return medium_dark_negative
            elif result < -0.07:
                return darkest_negative
            else:
                return grey
        elif result > 0:
            if result <= 0.01:
                return lightest_positive
            elif result <= 0.03:
                return medium_light_positive
            elif result <= 0.05:
                return medium_positive
            elif result <= 0.07:
                return medium_dark_positive
            elif result > 0.07:
                return darkest_positive
            else:
                return grey
        else:
            return grey

    def save(self, path_to_output):
        self.workbook.save(path_to_output)
        