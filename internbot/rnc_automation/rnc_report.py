from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font
from collections import OrderedDict
   
class RNCReport(object):

    def __init__ (self, models, state_name, path_to_topline=None):
        if path_to_topline != None:
            self.workbook = load_workbook(path_to_topline)
            first_round = False
        else:
            self.workbook = Workbook()
            first_round = True

        self.titles_style = Font( name='Calibri (Body)', 
                                size=11, 
                                bold=True, 
                                italic=False, 
                                vertAlign=None, 
                                underline='none', 
                                strike=False, 
                                color='000000')

        self.general_style = Font( name='Calibri (Body)', 
                                size=11, 
                                bold=False, 
                                italic=False, 
                                vertAlign=None, 
                                underline='none', 
                                strike=False, 
                                color='000000')
        
        self.net_style = Font( name='Calibri (Body)', 
                                size=11, 
                                bold=False, 
                                italic=True, 
                                vertAlign=None, 
                                underline='none', 
                                strike=False, 
                                color='000000')

        self.border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

        self.top_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'))

        self.middle_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'))

        self.bottom_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'),
                     bottom=Side(style='thick'))

        self.grey = PatternFill("solid", fgColor="B7B7B7")

        self.darkest_negative = PatternFill("solid", fgColor="B80001")
        self.medium_dark_negative = PatternFill("solid", fgColor="CD4748")
        self.medium_negative = PatternFill("solid", fgColor="DF8A8C")
        self.medium_light_negative = PatternFill("solid", fgColor="EAB9BB")
        self.lightest_negative = PatternFill("solid", fgColor="F1D2D5")

        self.darkest_positive = PatternFill("solid", fgColor="02A747")
        self.medium_dark_positive = PatternFill("solid", fgColor="2AB666")
        self.medium_positive = PatternFill("solid", fgColor="5AC88B")
        self.medium_light_positive = PatternFill("solid", fgColor="B6E7CE")
        self.lightest_positive = PatternFill("solid", fgColor="E6F6F0")

        self.models = models
        if first_round == True:
            if state_name == "Pennsylvania":
                self.abr = "PA"
            if state_name == "New Jersey":
                self.abr = "NJ"
            self.build_sheets()
            self.write_model()
            self.write_current_TOW()
            self.write_current_score()
            self.write_diff_previous()
            self.write_diff_first()
        else:
            self.check_for_new_models()
            self.insert_cols()
            self.write_current_TOW()
            self.write_current_score()
            self.update_diff_previous()
            self.update_diff_first()

    def build_sheets(self):
        sheet = self.workbook.get_sheet_by_name('Sheet')
        new_title = "Key-%s" % (self.abr)
        sheet.title = new_title
        new_sheet = self.abr
        self.workbook.create_sheet(str(new_sheet))
        
    def write_model(self):
        sheet = self.workbook.get_sheet_by_name(self.abr)
        sheet["A2"].value = "Model"
        sheet["A2"].font = self.titles_style
        sheet["A2"].border = self.border
        current_cell = 3
        model_names = self.models.list_names()
        for model_name in model_names:
            if model_name == "Turnout General":
                location = "A%s" % current_cell
                sheet[location].value = model_name
                sheet[location].font = self.titles_style
                sheet[location].border = self.border
                current_cell += 1
            else:
                location = "A%s" % current_cell
                sheet[location].value = model_name
                sheet[location].font = self.net_style
                sheet[location].border = self.border
                current_cell += 1
                variables = self.models.return_variables(model_name)
                use_middle = False
                top = True
                if len(variables) > 2:
                    use_middle = True
                for variable in variables:
                    if top == True:
                        location = "A%s" % current_cell
                        sheet[location].value = variable
                        sheet[location].font = self.general_style
                        sheet[location].border = self.top_border
                        current_cell += 1
                        top = False
                    elif use_middle == True:
                        use_middle = False
                        location = "A%s" % current_cell
                        sheet[location].value = variable
                        sheet[location].font = self.general_style
                        sheet[location].border = self.middle_border
                        current_cell += 1
                    else:
                        location = "A%s" % current_cell
                        sheet[location].value = variable
                        sheet[location].font = self.general_style
                        sheet[location].border = self.bottom_border
                        current_cell += 1

    def write_diff_previous(self):
        sheet = self.workbook.get_sheet_by_name(self.abr)
        sheet["B1"].value = "Current Round - Previous Round"
        sheet["B1"].font = self.titles_style
        sheet["B1"].border = self.border
        
        sheet["B2"].border = self.border
        
        current_cell = 3
        model_names = self.models.list_names()
        for model_name in model_names:
            if model_name == "Turnout General":
                location = "B%s" % current_cell
                sheet[location].fill = self.grey
                sheet[location].border = self.border
                current_cell += 1
            else:
                location = "B%s" % current_cell
                variables = self.models.return_variables(model_name)
                sheet[location].value = "--%"
                sheet[location].font = self.general_style
                sheet[location].border = self.border
                current_cell += 1
                use_middle = False
                top = True
                if len(variables) > 2:
                    use_middle = True
                for variable in variables:
                    if top == True:
                        location = "B%s" % current_cell
                        sheet[location].value = "--%"
                        sheet[location].font = self.general_style
                        sheet[location].border = self.top_border
                        current_cell += 1
                        top = False
                    elif use_middle == True:
                        use_middle = False
                        location = "B%s" % current_cell
                        sheet[location].value = "--%"
                        sheet[location].font = self.general_style
                        sheet[location].border = self.middle_border
                        current_cell += 1
                    else:
                        location = "B%s" % current_cell
                        sheet[location].value = "--%"
                        sheet[location].font = self.general_style
                        sheet[location].border = self.bottom_border
                        current_cell += 1

    def write_diff_first(self):
        sheet = self.workbook.get_sheet_by_name(self.abr)
        sheet["C1"].value = "Current Round - First Round"
        sheet["C1"].font = self.titles_style
        sheet["C1"].border = self.border
        
        sheet["C2"].border = self.border
        
        current_cell = 3
        model_names = self.models.list_names()
        for model_name in model_names:
            if model_name == "Turnout General":
                location = "C%s" % current_cell
                sheet[location].fill = self.grey
                sheet[location].border = self.border
                current_cell += 1
            else:
                location = "C%s" % current_cell
                variables = self.models.return_variables(model_name)
                sheet[location].value = "--%"
                sheet[location].font = self.general_style
                sheet[location].border = self.border
                current_cell += 1
                use_middle = False
                top = True
                if len(variables) > 2:
                    use_middle = True
                for variable in variables:
                    if top == True:
                        location = "C%s" % current_cell
                        sheet[location].value = "--%"
                        sheet[location].font = self.general_style
                        sheet[location].border = self.top_border
                        current_cell += 1
                        top = False
                    elif use_middle == True:
                        use_middle = False
                        location = "C%s" % current_cell
                        sheet[location].value = "--%"
                        sheet[location].font = self.general_style
                        sheet[location].border = self.middle_border
                        current_cell += 1
                    else:
                        location = "C%s" % current_cell
                        sheet[location].value = "--%"
                        sheet[location].font = self.general_style
                        sheet[location].border = self.bottom_border
                        current_cell += 1

    def write_current_TOW(self):
        sheet = self.workbook.get_sheet_by_name(self.abr)
        sheet["D2"].value = "Round 1 TOW"
        sheet["D2"].font = self.titles_style
        sheet["D2"].border = self.border
        
        current_cell = 3
        model_names = self.models.list_names()
        for model_name in model_names:
            if model_name == "Turnout General":
                location = "D%s" % current_cell
                sheet[location].fill = self.grey
                sheet[location].border = self.border
                current_cell += 1
            else:
                location = "D%s" % current_cell
                variables = self.models.return_variables(model_name)
                difference_count = len(variables)
                formula = "="
                while difference_count > 0:
                    formula_location = current_cell + difference_count
                    formula += "D%s" % formula_location
                    if difference_count > 1:
                        formula += " - "
                    difference_count -= 1
                sheet[location].value = formula
                sheet[location].number_format = '0%'
                sheet[location].font = self.general_style
                sheet[location].border = self.border
                current_cell += 1
                use_middle = False
                top = True
                if len(variables) > 2:
                    use_middle = True
                for variable in variables:
                    if top == True:
                        location = "D%s" % current_cell
                        sheet[location].value = variables[variable]
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.top_border
                        current_cell += 1
                        top = False
                    elif use_middle == True:
                        use_middle = False
                        location = "D%s" % current_cell
                        sheet[location].value = variables[variable]
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.middle_border
                        current_cell += 1
                    else:
                        location = "D%s" % current_cell
                        sheet[location].value = variables[variable]
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.bottom_border
                        current_cell += 1

    def write_current_score(self):
        sheet = self.workbook.get_sheet_by_name(self.abr)
        sheet["E2"].value = "Round 1 [Date]"
        sheet["E2"].font = self.titles_style
        sheet["E2"].border = self.border
        
        current_cell = 3
        model_names = self.models.list_names()
        for model_name in model_names:
            if model_name == "Turnout General":
                location = "E%s" % current_cell
                variables = self.models.return_variables(model_name)
                for variable in variables:
                    if variable == "Turnout General":
                        sheet[location].value = variables[variable]
                        sheet[location].number_format = '0%'
                sheet[location].border = self.border
                current_cell += 1
            else:
                location = "E%s" % current_cell
                variables = self.models.return_variables(model_name)
                difference_count = len(variables)
                formula = "="
                while difference_count > 0:
                    formula_location = current_cell + difference_count
                    formula += "E%s" % formula_location
                    if difference_count > 1:
                        formula += " - "
                    difference_count -= 1
                sheet[location].value = formula
                sheet[location].number_format = '0%'
                sheet[location].font = self.general_style
                sheet[location].border = self.border
                current_cell += 1
                use_middle = False
                top = True
                if len(variables) > 2:
                    use_middle = True
                for variable in variables:
                    if top == True:
                        location = "E%s" % current_cell
                        sheet[location].value = variables[variable]
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.top_border
                        current_cell += 1
                        top = False
                    elif use_middle == True:
                        use_middle = False
                        location = "E%s" % current_cell
                        sheet[location].value = variables[variable]
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.middle_border
                        current_cell += 1
                    else:
                        location = "E%s" % current_cell
                        sheet[location].value = variables[variable]
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.bottom_border
                        current_cell += 1

    def check_for_new_models(self):
        pass

    def insert_cols(self):
        pass

    def update_diff_previous(self):
        sheet = self.workbook.get_sheet_by_name(self.abr)
        sheet["B1"].value = "Current Round - Previous Round"
        sheet["B1"].font = self.titles_style
        sheet["B1"].border = self.border
        
        sheet["B2"].border = self.border
        
        current_cell = 3
        model_names = self.models.list_names()
        for model_name in model_names:
            if model_name == "Turnout General":
                location = "B%s" % current_cell
                sheet[location].fill = self.grey
                sheet[location].border = self.border
                current_cell += 1
            else:
                location = "B%s" % current_cell
                variables = self.models.return_variables(model_name)
                difference_count = len(variables)
                formula = "="
                while difference_count > 0:
                    formula_location = current_cell + difference_count
                    formula += "B%s" % formula_location
                    if difference_count > 1:
                        formula += " - "
                    difference_count -= 1
                sheet[location].value = formula
                self.highlight(sheet[location])
                sheet[location].number_format = '0%'
                sheet[location].font = self.general_style
                sheet[location].border = self.border
                current_cell += 1
                use_middle = False
                top = True
                if len(variables) > 2:
                    use_middle = True
                for variable in variables:
                    if top == True:
                        location = "B%s" % current_cell
                        sheet[location].value = variables[variable]
                        self.highlight(sheet[location])
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.top_border
                        current_cell += 1
                        top = False
                    elif use_middle == True:
                        use_middle = False
                        location = "B%s" % current_cell
                        sheet[location].value = variables[variable]
                        self.highlight(sheet[location])
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.middle_border
                        current_cell += 1
                    else:
                        location = "B%s" % current_cell
                        sheet[location].value = variables[variable]
                        self.highlight(sheet[location])
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.bottom_border
                        current_cell += 1

    def update_diff_first(self):
        sheet = self.workbook.get_sheet_by_name(self.abr)
        sheet["C1"].value = "Current Round - First Round"
        sheet["C1"].font = self.titles_style
        sheet["C1"].border = self.border
        
        sheet["C2"].border = self.border
        
        current_cell = 3
        model_names = self.models.list_names()
        for model_name in model_names:
            if model_name == "Turnout General":
                location = "C%s" % current_cell
                sheet[location].fill = self.grey
                sheet[location].border = self.border
                current_cell += 1
            else:
                location = "C%s" % current_cell
                variables = self.models.return_variables(model_name)
                difference_count = len(variables)
                formula = "="
                while difference_count > 0:
                    formula_location = current_cell + difference_count
                    formula += "C%s" % formula_location
                    if difference_count > 1:
                        formula += " - "
                    difference_count -= 1
                sheet[location].value = formula
                sheet[location].number_format = '0%'
                sheet[location].font = self.general_style
                sheet[location].border = self.border
                current_cell += 1
                use_middle = False
                top = True
                if len(variables) > 2:
                    use_middle = True
                for variable in variables:
                    if top == True:
                        location = "C%s" % current_cell
                        sheet[location].value = variables[variable]
                        self.highlight(sheet[location])
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.top_border
                        current_cell += 1
                        top = False
                    elif use_middle == True:
                        use_middle = False
                        location = "C%s" % current_cell
                        sheet[location].value = variables[variable]
                        self.highlight(sheet[location])
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.middle_border
                        current_cell += 1
                    else:
                        location = "C%s" % current_cell
                        sheet[location].value = variables[variable]
                        self.highlight(sheet[location])
                        sheet[location].number_format = '0%'
                        sheet[location].font = self.general_style
                        sheet[location].border = self.bottom_border
                        current_cell += 1

    def highlight(self, cell):
        if cell.value < 0:
            if cell.value >= -0.01:
                cell.fill = self.lightest_negative
            elif cell.value >= -0.03:
                cell.fill = self.medium_light_negative
            elif cell.value >= -0.05:
                cell.fill = self.medium_negative
            elif cell.value >= -0.07:
                cell.fill = self.medium_dark_negative
            else:
                cell.fill = self.darkest_negative
        elif cell.value > 0:
            if cell.value <= 0.01:
                cell.fill = self.lightest_positive
            elif cell.value <= 0.03:
                cell.fill = self.medium_light_positive
            elif cell.value <= 0.05:
                cell.fill = self.medium_positive
            elif cell.value <= 0.07:
                cell.fill = self.medium_dark_positive
            else:
                cell.fill = self.darkest_positive

    def save(self, path_to_output):
        self.workbook.save(path_to_output)
        