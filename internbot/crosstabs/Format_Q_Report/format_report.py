from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
from collections import OrderedDict

class QParser(object):

    def __init__(self, path_to_workbook, is_qualtrics=True):
        print "Loading workbook"
        self.__workbook = load_workbook(path_to_workbook)
        self.__is_qualtrics = is_qualtrics
        self.__tables = OrderedDict()

        # fill colors
        if is_qualtrics:
            self.__header_fill = PatternFill("solid", fgColor = "1E262E")
            self.__hi_significant_fill = PatternFill("solid", fgColor = "2DCCD3")
            self.__lo_significant_fill = PatternFill("solid", fgColor = "2DCCD3")
        else:
            self.__header_fill = PatternFill("solid", fgColor = "0F243E")
            self.__hi_significant_fill = PatternFill("solid", fgColor = "2083E7")
            self.__lo_significant_fill = PatternFill("solid", fgColor = "2083E7")
        self.__table_fill = PatternFill("solid", fgColor = "E7E6E6")
        self.__white_fill = PatternFill("solid", fgColor = "FFFFFF")

        # font styles
        self.__font_reg = Font(name = 'Arial', size = 8)
        self.__font_small = Font(name = 'Arial', size = 7)
        self.__font_bold = Font(name = 'Arial', size = 8, bold = True)
        self.__font_back_hyperlink = Font(name = 'Arial', size = 8, color = "A2AAAD")
        self.__font_toc_table = Font(name = 'Arial', size = 8, underline = 'single')

        # alignments
        self.__align_banners = Alignment(horizontal = "center", vertical = "bottom", wrapText = True)
        self.__align_names = Alignment(horizontal = "left", vertical = "center", wrapText = True)
        self.__align_center = Alignment(horizontal = "center", vertical = "center", wrapText = True)
        self.__align_left = Alignment(horizontal = "left", vertical = "center", wrapText = True)

        # borders
        self.__thin_bottom = Border(bottom = Side(style = 'thin'))
        self.__thick_top = Border(top = Side(style = 'thick'))
        self.__thin_top = Border(top = Side(style = 'thin'))
        self.__thick_left = Border(left = Side(style = 'thick'))

        # calculate the excel alphabet from A to ZZZ
        alphabet = []
        for letter in range(65, 91):
            alphabet.append(chr(letter))

        self.__extend_alphabet = []
        self.__extend_alphabet.extend(alphabet)
        full_alphabet = []
        double_alphabet = []
        for letter in range(65, 91):
            full_alphabet.append(chr(letter))
        index = 0
        while index < len(full_alphabet):
            for letter in full_alphabet:
                double_letters = "%s%s" % (full_alphabet[index], letter)
                double_alphabet.append(double_letters)
                self.__extend_alphabet.append(double_letters)
            index += 1
        index = 0
        while index < len(double_alphabet):
            for letter in full_alphabet:
                triple_letters = "%s%s" % (double_alphabet[index], letter)
                self.__extend_alphabet.append(triple_letters)
            index += 1    

    def format_report(self):
        sheet_no = 1
        for sheet in self.__workbook.worksheets:
            if sheet.title != 'TOC':
                self.format_sheet(sheet, sheet_no)
                sheet_no += 1

        try:
            toc_sheet = self.__workbook.get_sheet_by_name("TOC")
            self.__workbook.remove_sheet(toc_sheet)
        except KeyError:
            pass

        print "Done!"
        return self.__tables

    def create_toc(self, tables):
        sheet = self.__workbook.create_sheet("TOC", 0)
        if not self.__is_qualtrics:
            sheet.row_dimensions[1].height = 52
        else:
            sheet.row_dimensions[1].height = 35
        
        sheet.column_dimensions["A"].width = 9
        sheet.column_dimensions["B"].width = 100
        sheet.column_dimensions["C"].width = 33
        sheet.column_dimensions["D"].width = 29.5

        sheet["A1"].fill = self.__header_fill
        sheet["B1"].fill = self.__header_fill
        sheet["C1"].fill = self.__header_fill
        sheet["D1"].fill = self.__header_fill

        if self.__is_qualtrics:
            logo = Image("templates_images/QLogo.png")
        else:
            logo = Image("templates_images/y2_xtabs.png")
        sheet.add_image(logo, "D1")

        sheet["A2"].font = self.__font_bold
        sheet["A2"].value = "Table #"
        sheet["A2"].alignment = self.__align_center

        sheet["B2"].font = self.__font_bold
        sheet["B2"].value = "Question Title"
        sheet["B2"].alignment = self.__align_center

        sheet["C2"].font = self.__font_bold
        sheet["C2"].value = "Base Description"
        sheet["C2"].alignment = self.__align_center

        sheet["D2"].font = self.__font_bold
        sheet["D2"].value = "Base Size (N count)"
        sheet["D2"].alignment = self.__align_center

        current_row = 3
        for key, value in tables.iteritems():
            table_no_cell = "A%s" % str(current_row)
            question_title_cell = "B%s" % str(current_row)
            base_desc_cell = "C%s" % str(current_row)
            base_size_cell = "D%s" % str(current_row)

            sheet[table_no_cell].font = self.__font_toc_table
            if int(value.name) < 10:
                table_name = "Table 0%s" % value.name
            else:
                table_name = "Table %s" % value.name
            sheet[table_no_cell].value = "=HYPERLINK(\"#'%s'!A1\",\"%s\")" % (table_name, table_name)
            sheet[table_no_cell].alignment = self.__align_center
            sheet[table_no_cell].border = self.__thin_bottom

            sheet[question_title_cell].value = value.prompt
            sheet[question_title_cell].font = self.__font_reg
            sheet[question_title_cell].alignment = self.__align_left
            sheet[question_title_cell].border = self.__thin_bottom

            sheet[base_desc_cell].value = value.description
            sheet[base_desc_cell].font = self.__font_reg
            sheet[base_desc_cell].alignment = self.__align_left
            sheet[base_desc_cell].border = self.__thin_bottom

            sheet[base_size_cell].value = value.size
            sheet[base_size_cell].font = self.__font_reg
            sheet[base_size_cell].alignment = self.__align_left
            sheet[base_size_cell].border = self.__thin_bottom

            current_row += 1

    def format_sheet(self, sheet, sheet_no):
        self.__col_names = []
        if sheet_no < 10:
            sheet_title = "Table 0%s" % str(sheet_no)
        else:
            sheet_title = "Table %s" % str(sheet_no)
        sheet.title = sheet_title

        print "Formatting: %s" % sheet_title

        table = TOCTable(sheet_no)
        start_table_row = self.parse_col_names(sheet)
        self.format_hyperlink(sheet)
        table = self.format_titles(sheet, table)
        self.format_banners(sheet, start_table_row)
        self.format_responses(sheet, start_table_row)
        self.format_table_contents(sheet, start_table_row)
        self.format_stats_def(sheet)
        self.add_table_borders(sheet)
        self.__tables[sheet_no] = table

        # merge top left table corner
        sheet.merge_cells(start_row=3, start_column=1, end_row=start_table_row-1, end_column=1)
        if not self.__is_qualtrics:
            sheet.row_dimensions[1].height = 52

        # there's a bug with insert cols (and probably insert rows) where the code doesn't preserve merged cells
        # we need to un merge cells in span if there are any
        #sheet.unmerge_cells('C3:E3') 
        #sheet.unmerge_cells('F3:L3')
        #sheet.insert_cols(1)
        #sheet.move_range("A1:L19", cols=1)
        #sheet.insert_cols(1)

    def parse_col_names(self, sheet):
        sheet.column_dimensions["A"].width = 25.5
        column_names_row = 1

        # split up merged cells in column A to prevent corruption in excel file
        for cell in sheet["A"]:
            if "MergedCell" in str(type(cell)):
                range = "%s%s:%s" % (self.__extend_alphabet[cell.column-1], str(cell.row-1), cell.coordinate)
                sheet.unmerge_cells(range)

            if cell.value == "Column Names":
                break
            else:
                column_names_row += 1

        # find row of column names: grab the column names
        for cell in sheet[column_names_row]:
            if cell.value is None:
                break
            elif cell.value != "Column Names":
                self.__col_names.append(cell.value)
        sheet.delete_rows(column_names_row, 1)

        # the row after column names describes the stats of the table
        self.__stats_desc = column_names_row + 1

        # figure out where the crosstab calculations start
        # insert a row before that for column names
        start_table_row = 1
        for cell in sheet["B"]:
            if cell.number_format != "General":
                break
            else:
                start_table_row += 1
        sheet.insert_rows(start_table_row - 1)

        # populate newly inserted row
        sheet.row_dimensions[start_table_row-1].height = 16
        col_index = 1
        for name in self.__col_names:
            col_letter = self.__extend_alphabet[col_index]
            current_cell = "%s%s" % (col_letter, str(start_table_row - 1))
            sheet[current_cell].value = "[%s]" % name
            sheet[current_cell].font = self.__font_reg
            sheet[current_cell].alignment = self.__align_center
            col_index += 1

        return start_table_row + 1

    def format_hyperlink(self, sheet):
        # check if we need to add a row
        if sheet["A1"].value != "Back to TOC":
            sheet.insert_rows(1)
        else:
            sheet["A1"].value = ""

        # back to TOC cell
        sheet.row_dimensions[1].height = 35
        sheet.row_dimensions[2].height = 36
        sheet["A3"].value = ""
        sheet["A1"].value = '=HYPERLINK("#TOC!A1","Return to Table of Contents")'
        sheet["A1"].font = self.__font_back_hyperlink
        sheet["A1"].alignment = self.__align_left

    def format_titles(self, sheet, table):
        current_row = 1
        current_col = 0

        # add header and table fills based off the size of the table (easily calculated from column names)
        for col in range(len(self.__col_names) + 1):
            col_letter = self.__extend_alphabet[col]
            current_cell = "%s%s" % (col_letter, str(current_row))
            sheet[current_cell].fill = self.__header_fill
            next_row_cell = "%s%s" % (col_letter, str(current_row+1))
            sheet[next_row_cell].fill = self.__table_fill
            current_col += 1

        # merge that top row for aesthetics
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=current_col - 3)

        # add logos
        current_cell = "%s%s" % (self.__extend_alphabet[current_col - 3], str(current_row))
        if self.__is_qualtrics:
            logo = Image("templates_images/QLogo.png")
        else:
            logo = Image("templates_images/y2_xtabs.png")
        sheet.add_image(logo, current_cell)

        # figure out where the base description of table will be added
        next_row_cell = "%s%s" % (self.__extend_alphabet[current_col - 3], str(current_row + 1))
        sheet[next_row_cell].font = self.__font_reg
        sheet[next_row_cell].alignment = self.__align_names

        table.location = "%s%s" % (self.__extend_alphabet[current_col-3], str(2))

        # merge table title and base description cells
        sheet.merge_cells(start_column=1, end_column=current_col-3, start_row=2, end_row=2)
        sheet.merge_cells(start_column=current_col-2, end_column=current_col, start_row=2, end_row=2)

        # add table title
        sheet["A2"].font = self.__font_reg
        sheet["A2"].alignment = self.__align_names
        table_title = sheet["A2"].value
        names = table_title.split("by")
        del names[-1]
        title = ""
        for name in names:
            title = title + name
        final_title = "%s - %s" % (sheet.title, title)
        sheet["A2"].value = final_title

        table.prompt = title
        return table

    def format_banners(self, sheet, start_row):
        current_row = start_row-1
        sheet.row_dimensions[current_row].height = 49
        while current_row > 1:
            col_index = 1
            while col_index < len(self.__col_names) + 1:
                current_cell = "%s%s" % (self.__extend_alphabet[col_index], str(current_row))
                sheet[current_cell].font = self.__font_reg
                sheet[current_cell].alignment = self.__align_banners
                col_index += 1
            current_row -= 1

    def format_responses(self, sheet, start_row):
        self.__border_rows = []
        merge_cells = []
        is_top = True
        while start_row < self.__stats_desc:
            current_cell = "A%s" % str(start_row)
            if sheet[current_cell].value == "NET":
                sheet[current_cell].value = "Total"

            # merge response cells together
            if sheet[current_cell].value is not None:
                self.__border_rows.append(start_row)
                if is_top:
                    is_top = False
                else:
                    range = "%s:%s" % (merge_cells[0], merge_cells[len(merge_cells)-1])
                    merge_cells = []
                    sheet.merge_cells(range)
                merge_cells.append(sheet[current_cell].coordinate)
            else:
                merge_cells.append(sheet[current_cell].coordinate)

            sheet[current_cell].font = self.__font_bold
            sheet[current_cell].alignment = self.__align_left
            sheet[current_cell].fill = self.__table_fill
            start_row += 1

        range = "%s:%s" % (merge_cells[0], merge_cells[len(merge_cells)-1])
        merge_cells = []
        sheet.merge_cells(range)

    def format_table_contents(self, sheet, start_row):
        current_row = start_row

        # change font/alignment and highlight significant cells
        while current_row < self.__stats_desc:
            col_no = 1
            while col_no < len(self.__col_names) + 1:
                current_cell = "%s%s" % (self.__extend_alphabet[col_no], str(current_row))
                sheet[current_cell].font = self.__font_reg
                sheet[current_cell].alignment = self.__align_center
                sheet[current_cell].fill = self.__white_fill
                if sheet[current_cell].data_type == 's':
                    if sheet[current_cell].value.isupper():
                        sheet[current_cell].fill = self.__hi_significant_fill
                    elif sheet[current_cell].value.islower():
                        sheet[current_cell].fill = self.__lo_significant_fill
                col_no += 1
            current_row += 1

    def format_stats_def(self, sheet):
        current_row = self.__stats_desc
        while True:
            current_cell = "A%s" % str(current_row)
            if sheet[current_cell].value is not None:
                sheet[current_cell].font = self.__font_small
            else:
                break
            current_row += 1

    def add_table_borders(self, sheet):
        for row_no in self.__border_rows:
            col_no = 0
            while col_no < len(self.__col_names) + 1:
                current_cell = "%s%s" % (self.__extend_alphabet[col_no], str(row_no))
                sheet[current_cell].border = self.__thin_top
                col_no += 1

        col_no = 0
        while col_no < len(self.__col_names) + 1:
            current_cell = "%s%s" % (self.__extend_alphabet[col_no], str(self.__stats_desc))
            sheet[current_cell].border = self.__thick_top
            col_no += 1

        current_row = 1
        while current_row < self.__stats_desc:
            current_cell = "%s%s" % (self.__extend_alphabet[col_no], str(current_row))
            sheet[current_cell].border = self.__thick_left
            current_row += 1

    def add_bases(self, tables):
        for key, value in tables.iteritems():
            if int(value.name) < 10:
                table_name = "Table 0%s" % value.name
            else:
                table_name = "Table %s" % value.name
            
            sheet = self.__workbook.get_sheet_by_name(table_name)

            sheet[value.location].value = value.description
            sheet[value.location].font = self.__font_reg
            sheet[value.location].alignment = self.__align_names

        self.create_toc(tables)

    def save(self, path_to_output):
        self.__workbook.save(path_to_output)

class TOCTable(object):

    def __init__(self, table_name):
        self.__name = str(table_name)

    @property
    def prompt(self):
        return self.__prompt

    @property
    def description(self):
        return self.__description

    @property
    def size(self):
        return self.__size

    @property
    def name(self):
        return self.__name

    @property
    def location(self):
        return self.__location

    @prompt.setter
    def prompt(self, prompt):
        self.__prompt = prompt

    @description.setter
    def description(self, desc):
        self.__description = desc

    @size.setter
    def size(self, pop):
        self.__size = pop

    @location.setter
    def location(self, location):
        self.__location = location

