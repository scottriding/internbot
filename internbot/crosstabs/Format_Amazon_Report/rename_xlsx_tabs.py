from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
import csv

class RenameTabs(object):

    def __init__(self):
        self.__sheet_names = []
        self.__white_highlight = PatternFill("solid", fgColor="FFFFFF")
        self.__yellow_highlight = PatternFill("solid", fgColor="FFFF00")
        self.__font_reg = Font(name = 'Verdana', size = 8)
        self.__font_title = Font(name = "Calibri (Body)", size=11)
        self.__font_title_bold = Font(name = 'Verdana', size = 8, bold = True)
        self.__font_hyper = Font(name="Arial", size=10, color="0563C1", underline = 'single')
        self.__center_align = Alignment(horizontal="center", vertical="center")
        self.__center_top = Alignment(horizontal="center", vertical="top")
        self.__all_align = Alignment(horizontal="left", vertical="top", wrapText = True)
        self.__thick_sides = Border(right=Side(style='thick'), left=Side(style='thick'))
        self.__thick_top = Border(top=Side(style='thick'))
        self.__thick_all = Border(top=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thick'), left=Side(style='thick'))

        # calculate the excel alphabet from A to ZZZ
        alphabet = []
        for letter in range(65, 91):
            alphabet.append(chr(letter))

        self.extend_alphabet = []
        self.extend_alphabet.extend(alphabet)
        full_alphabet = []
        double_alphabet = []
        for letter in range(65, 91):
            full_alphabet.append(chr(letter))
        index = 0
        while index < len(full_alphabet):
            for letter in full_alphabet:
                double_letters = "%s%s" % (full_alphabet[index], letter)
                double_alphabet.append(double_letters)
                self.extend_alphabet.append(double_letters)
            index += 1
        index = 0
        while index < len(double_alphabet):
            for letter in full_alphabet:
                triple_letters = "%s%s" % (double_alphabet[index], letter)
                self.extend_alphabet.append(triple_letters)
            index += 1

    def rename(self, path_to_xlsx, path_to_toc, path_to_output):
        print("Reading in workbook")
        output_workbook = load_workbook(path_to_xlsx)
        try:
            toc_sheet = output_workbook.create_sheet("TOC", 0)
            self.create_toc(path_to_toc, toc_sheet)
        except:
            toc_sheet = output_workbook["TOC"]
            self.write_table_of_contents(path_to_toc, toc_sheet)
        self.rename_worksheets(output_workbook)
        return output_workbook

    def create_toc(self, path_to_tables, toc_sheet):
        self.write_titles(toc_sheet)
        self.write_table_of_contents(path_to_tables, toc_sheet)
        hyperlink_col = 'B'
        table_index = 'C'
        current_row = 10 
        check_cell = "%s%s" % (table_index, current_row)
        while toc_sheet[check_cell].value is not None:
            hyperlink_form = "=HYPERLINK(CONCATENATE(\"[\",$B$56,\"]\",D%s,\"!A1\"), C%s)" % (current_row, current_row)
            hyperlink_cell = "%s%s" % (hyperlink_col, current_row)
            toc_sheet[hyperlink_cell].value = hyperlink_form
            toc_sheet[hyperlink_cell].fill = self.__white_highlight
            toc_sheet[hyperlink_cell].font = self.__font_hyper
            toc_sheet[hyperlink_cell].border = self.__thick_sides
            toc_sheet[hyperlink_cell].alignment = self.__center_top
            current_row += 1
            check_cell = "%s%s" % (table_index, current_row)

    def write_titles(self, toc_sheet):
        toc_sheet.column_dimensions["A"].width = 9
        toc_sheet.column_dimensions["B"].width = 9.83
        toc_sheet.column_dimensions["C"].width = 2.17
        toc_sheet.column_dimensions["D"].width = 17.17
        toc_sheet.column_dimensions["E"].width = 100
        toc_sheet.column_dimensions["F"].width = 33

        toc_sheet["B9"].value = "Table No." 
        toc_sheet["B9"].font = self.__font_title_bold
        toc_sheet["B9"].alignment = self.__center_align
        toc_sheet["B9"].border = self.__thick_all

        toc_sheet["C9"].value = "#"
        toc_sheet["C9"].font = self.__font_title_bold
        toc_sheet["C9"].alignment = self.__center_align
        toc_sheet["C9"].border = self.__thick_all

        toc_sheet["D9"].border = self.__thick_all

        toc_sheet["E9"].value = "Question Title"
        toc_sheet["E9"].font = self.__font_title_bold
        toc_sheet["E9"].alignment = self.__center_align
        toc_sheet["E9"].border = self.__thick_all

        toc_sheet["F9"].value = "Base"
        toc_sheet["F9"].font = self.__font_title_bold
        toc_sheet["F9"].alignment = self.__center_align
        toc_sheet["F9"].border = self.__thick_all

        toc_sheet["E8"].value = "CX Quality Survey Time Series Wave #s - All Respondents"
        toc_sheet["E8"].font = self.__font_title_bold
        toc_sheet["E8"].alignment = self.__center_align
        toc_sheet["E8"].fill = self.__yellow_highlight

        toc_sheet.column_dimensions['C'].hidden= True

        logo = Image("templates_images/Old_QLogo.png")
        toc_sheet.add_image(logo, "B2")

    def write_table_of_contents(self, path_to_tables, toc_sheet):
        print("Writing TOC")
        table_index = 'C'
        question_name = 'D'
        question_prompt = 'E'
        base = 'F'
        iteration = 10
        with open(path_to_tables, 'r') as csvfile:
            file = csv.DictReader(csvfile, quotechar = '"')
            for table_info in file:
                toc_sheet[table_index + str(iteration)].value = table_info['TableIndex']
                toc_sheet[table_index + str(iteration)].fill = self.__white_highlight

                sheet_name = table_info['VariableName'].translate(None,"$")
                toc_sheet[question_name + str(iteration)].value = sheet_name
                toc_sheet[question_name + str(iteration)].fill = self.__white_highlight
                toc_sheet[question_name + str(iteration)].font = self.__font_reg
                toc_sheet[question_name + str(iteration)].border = self.__thick_sides
                toc_sheet[question_name + str(iteration)].alignment = self.__all_align

                toc_sheet[question_prompt + str(iteration)].value = table_info['Title']
                toc_sheet[question_prompt + str(iteration)].fill = self.__white_highlight
                toc_sheet[question_prompt + str(iteration)].font = self.__font_title
                toc_sheet[question_prompt + str(iteration)].border = self.__thick_sides
                toc_sheet[question_prompt + str(iteration)].alignment = self.__all_align

                toc_sheet[base + str(iteration)].value = table_info['Base']
                toc_sheet[base + str(iteration)].fill = self.__white_highlight
                toc_sheet[base + str(iteration)].font = self.__font_title
                toc_sheet[base + str(iteration)].border = self.__thick_sides
                toc_sheet[base + str(iteration)].alignment = self.__all_align
                iteration += 1
                self.__sheet_names.append(sheet_name)
        
        toc_sheet[table_index + str(iteration)].border = self.__thick_top
        toc_sheet[question_name + str(iteration)].border = self.__thick_top
        toc_sheet[question_prompt + str(iteration)].border = self.__thick_top
        toc_sheet[base + str(iteration)].border = self.__thick_top
        toc_sheet["B" + str(iteration)].border = self.__thick_top

        toc_sheet["B" + str(iteration + 1)].value = "=MID(CELL(\"filename\",A1),FIND(\"[\",CELL(\"filename\",A1))+1,FIND(\"]\", CELL(\"filename\",A1))-FIND(\"[\",CELL(\"filename\",A1))-1)"
        toc_sheet["B" + str(iteration + 1)].font = self.__font_reg

    def rename_worksheets(self, workbook):
        print("Naming worksheets")
        index = 0
        for sheet in workbook.worksheets:
            if sheet.title != 'TOC':
                sheet.title = self.__sheet_names[index]
                index += 1