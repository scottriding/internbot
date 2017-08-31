from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from cell import Cells, Cell, FrequencyCell, PopulationCell, SignificantMarker

class Highlighter(object):

    def __init__(self, path_to_xlsx):
        self.workbook = load_workbook(path_to_xlsx)
        self.groups = {}
        self.responses = {}
        self.__cells = Cells()
        self.highlight_style = PatternFill("solid", fgColor="DDDDDD")
    
    def highlight(self):
        for sheet in self.workbook.worksheets:
            if sheet.title == 'Country':
                self.parse_rows(sheet)
                self.parse_columns(sheet)
                self.create_cells()
                self.assign_significant(sheet)
                self.highlight_significant(sheet)
        self.workbook.save('/Users/y2analytics/Desktop/Highlighted.xlsx')
        
    def parse_rows(self, sheet):
        response_column = sheet['B']
        for cell in response_column:
            if cell.value is not None and \
               cell.value != 'Sigma' and \
               cell.value not in self.responses.keys():
                self.responses[cell.value] = cell.row
            elif cell.value is not None and \
                cell.value in self.responses.keys():
                temp_storage = self.responses.get(cell.value)
                self.responses[cell.value] = [temp_storage]
                self.responses[cell.value].append(cell.row)

    def parse_columns(self, sheet):
        group_row = sheet['3']
        for cell in group_row:
            if cell.value is not None and \
               cell.value not in self.groups.keys():
                self.groups[cell.value] = cell.column

    def create_cells(self):
        for group in self.groups:
            for label in self.responses:
                location1 = str(self.groups[group]) + str(self.responses[label][0])
                location2 = str(self.groups[group]) + str(self.responses[label][0] + 1)
                location3 = str(self.groups[group]) + str(self.responses[label][1])
                self.__cells.add(PopulationCell(label, group, location1))
                self.__cells.add(FrequencyCell(label, group, location2))
                self.__cells.add(SignificantMarker(label, group, location3))

    def assign_significant(self, sheet):
        for cell in self.__cells:
            if cell.type == 'SignificantMarker':
                sheet_cell = sheet[cell.location].value
                if sheet_cell is not None and sheet_cell.isupper() is True:
                    match_cell = self.__cells.matching_cells(cell)
                    match_cell.is_significant = True

    def highlight_significant(self, sheet):
        for cell in self.__cells:
            if cell.is_significant is True:
                sheet[cell.location].fill = self.highlight_style
        
highlighter = Highlighter('/Users/y2analytics/Documents/1.xlsx')
highlighter.highlight()