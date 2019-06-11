from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
import os

class CrosstabReportWriter(object):

    def __init__ (self, tables, resources_filepath):
        self.__tables = tables
        self.resources_filepath = resources_filepath
        self.__workbook = Workbook()

        self.highlight_style = PatternFill("solid", fgColor="2DCCD3")
        self.shading_style = PatternFill("solid", fgColor="E7E6E6")

        self.__font_reg = Font(name = 'Arial', size = 8)
        self.__font_bold = Font(name = 'Arial', size = 8, bold = True)
        self.__center_align = Alignment(horizontal="left", vertical="center")
        self.__align = Alignment(horizontal="center", vertical="center")

        self.__report_header = PatternFill("solid", fgColor="1E262E")
        self.__table_liner = PatternFill("solid", fgColor="E7E6E6")

        self.__thin_bottom = Border(bottom=Side(style='thin'))
        self.__thick_bottom = Border(bottom=Side(style='thick'))
        self.__thin_top = Border(top=Side(style='thin'))
        self.__thin_left = Border(left=Side(style='thin'))

        # calculate the excel alphabet from A to ZZZ
        alphabet = []
        for letter in range(65, 91):
            alphabet.append(chr(letter))

        self.extend_alphabet = []
        self.extend_alphabet.extend(alphabet)
        full_alphabet = []
        double_alphabet = []
        for letter in range(65, 91):
            full_alphabet.append(chr(letter))
        index = 0
        while index < len(full_alphabet):
            for letter in full_alphabet:
                double_letters = "%s%s" % (full_alphabet[index], letter)
                double_alphabet.append(double_letters)
                self.extend_alphabet.append(double_letters)
            index += 1
        index = 0
        while index < len(double_alphabet):
            for letter in full_alphabet:
                triple_letters = "%s%s" % (double_alphabet[index], letter)
                self.extend_alphabet.append(triple_letters)
            index += 1              

    def save(self, path_to_output):
        self.__workbook.save(path_to_output)

    def write_report(self):
        if self.__workbook.get_sheet_by_name("Sheet") is not None:
            toc_sheet = self.__workbook.get_sheet_by_name('Sheet')
            toc_sheet.title = "TOC"
        else:
            toc_sheet = self.__workbook.create_sheet("TOC")

        self.write_toc(toc_sheet)

        iteration = 1
        for table in self.__tables:
            if iteration < 10:
                table_name = "Table 0%s" % iteration
            else:
                table_name = "Table %s" % iteration

            new_sheet = self.__workbook.create_sheet(table_name)
            self.write_table(new_sheet, table)
            iteration += 1
        print("Finished!")

    def write_toc(self, sheet):
        print("Writing table of contents")
        self.write_toc_titles(sheet)
        current_row = 3
        iteration = 1

        for table in self.__tables:

            current_table_cell = "A%s" % current_row
            current_question_title = "B%s" % current_row
            current_base_desc = "C%s" % current_row
            current_base_size = "D%s" % current_row

            if iteration < 10:
                current_table_name = "Table 0%s" % iteration
            else:
                current_table_name = "Table %s" % iteration

            sheet[current_table_cell].font = Font(name = 'Arial', size = 8, underline = 'single')
            sheet[current_table_cell].value = "=HYPERLINK(\"#'%s'!A1\",\"%s\")" % (current_table_name, current_table_name)
            sheet[current_table_cell].alignment = self.__center_align
            sheet[current_table_cell].border = self.__thin_bottom

            current_name = table.name[11:]
            current_name = current_name.replace("$", "")
            sheet[current_question_title].value = current_name
            sheet[current_question_title].font = self.__font_reg
            sheet[current_question_title].alignment = self.__center_align
            sheet[current_question_title].border = self.__thin_bottom

            sheet[current_base_desc].font = self.__font_reg
            sheet[current_base_desc].value = table.base_description
            sheet[current_base_desc].alignment = self.__center_align
            sheet[current_base_desc].border = self.__thin_bottom

            sheet[current_base_size].font = self.__font_reg
            sheet[current_base_size].value = table.base_size
            sheet[current_base_size].alignment = self.__center_align
            sheet[current_base_size].border = self.__thin_bottom

            iteration += 1
            current_row += 1            
        
    def write_toc_titles(self, sheet):
        sheet.row_dimensions[1].height = 35
        sheet.column_dimensions["A"].width = 9
        sheet.column_dimensions["B"].width = 61.5
        sheet.column_dimensions["C"].width = 13.33
        sheet.column_dimensions["D"].width = 29.5

        sheet["A1"].fill = self.__report_header
        sheet["B1"].fill = self.__report_header
        sheet["C1"].fill = self.__report_header
        sheet["D1"].fill = self.__report_header

        image = Image(os.path.join(self.resources_filepath, "QLogo.png"))
        sheet.add_image(image, "D1")

        sheet["A2"].font = self.__font_bold
        sheet["A2"].value = "Table #"
        sheet["A2"].alignment = self.__center_align

        sheet["B2"].font = self.__font_bold
        sheet["B2"].value = "Question Title"
        sheet["B2"].alignment = self.__center_align

        sheet["C2"].font = self.__font_bold
        sheet["C2"].value = "Base Description"
        sheet["C2"].alignment = self.__center_align

        sheet["D2"].font = self.__font_bold
        sheet["D2"].value = "Base Size (N count)"
        sheet["D2"].alignment = self.__center_align

    def write_table(self, sheet, table):
        to_print = "Writing: %s" % sheet.title
        print(to_print)
        self.write_table_titles(sheet, table)
        current_row = self.write_banners(sheet, table)
        self.write_reponse_details(sheet, table, current_row)
        current_row = self.write_cross_details(sheet, table, current_row)
        self.sig_definition(sheet, current_row, table)
        self.place_table_border(sheet, current_row, table)
        
    def write_table_titles(self, sheet, table):
        number_of_cols = table.count_banner_pts + 3
        sheet.row_dimensions[1].height = 35
        sheet.row_dimensions[2].height = 36

        sheet["A1"].fill = self.__report_header
        sheet["A1"].value = '=HYPERLINK("#TOC!A1","Return to Table of Contents")'
        sheet["A1"].font = Font(name = 'Arial', size = 8, color = "A2AAAD")
        sheet["A1"].alignment = self.__center_align

        sheet["A2"].fill = self.__table_liner
        current_name = table.name.replace("$", "")
        sheet["A2"].value = current_name
        sheet["A2"].font = self.__font_reg
        sheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        sheet["B1"].fill = self.__report_header
        sheet["B2"].fill = self.__table_liner

        sheet["C1"].fill = self.__report_header
        sheet["C2"].fill = self.__table_liner

        current_cell = "D1"
        index = 0
        current_row = 1
        
        while index < number_of_cols:
            current_col = "%s" % self.extend_alphabet[index]
            sheet.column_dimensions[current_col].width = 15

            current_cell = "%s%s" % (self.extend_alphabet[index], current_row)
            sheet[current_cell].fill = self.__report_header

            next_row_cell = "%s%s" % (self.extend_alphabet[index], current_row + 1)
            sheet[next_row_cell].fill = self.__table_liner

            index += 1

        current_cell = "%s%s" % (self.extend_alphabet[index - 3], current_row)
        logo = Image(os.path.join(self.resources_filepath, "QLogo.png"))
        sheet.add_image(logo, current_cell)

        next_row_cell = "%s%s" % (self.extend_alphabet[index - 3], current_row + 1)
        sheet[next_row_cell].value = table.base_description
        sheet[next_row_cell].font = self.__font_reg
        sheet[next_row_cell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        sheet.merge_cells(start_column=1, end_column=index-3, start_row=2, end_row=2)
        sheet.merge_cells(start_column=index-2, end_column=index, start_row=2, end_row=2)

    def write_banners(self, sheet, table):
        col_index = 3
        letter_index = 1
        letter_row = 0

        for banner in table.banners:
            current_row = 3
            sheet.row_dimensions[current_row].height = 34
            banner_depth = len(banner)
            letter_row = 2 + banner_depth
            banner_index = 0
            while banner_index < banner_depth:
                if current_row == letter_row:
                    letter_cell = "%s%s" % (self.extend_alphabet[col_index], str(current_row))
                    letter_value = "[%s]" % self.extend_alphabet[letter_index]
                    sheet[letter_cell].value = letter_value
                    sheet[letter_cell].font = self.__font_bold
                    sheet[letter_cell].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                    letter_index += 1
                else:
                    banner_cell = "%s%s" % (self.extend_alphabet[col_index], str(current_row))
                    sheet[banner_cell].value = banner[banner_index]
                    sheet[banner_cell].font = self.__font_bold
                    sheet[banner_cell].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                    banner_index += 1
                current_row += 1
            col_index += 1        

        self.merge_banners(sheet, table, letter_row)
        sheet.merge_cells(start_column=1, end_column=2, start_row=3, end_row=current_row-1)
        letter_cell = "C%s" % str(letter_row)
        average_cell = "C%s" % str(letter_row + 1)
        sheet[letter_cell].value = "[A]"
        sheet[letter_cell].font = self.__font_bold
        sheet[letter_cell].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        sheet[average_cell].value = "Average"
        sheet[average_cell].font = self.__font_bold
        sheet[average_cell].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        return current_row

    def merge_banners(self, sheet, table, letter_row):
        current_row = 3
        
        to_merge = {}

        while current_row < letter_row:
            col_index = 3
            next_cell = "%s%s" % (self.extend_alphabet[col_index + 1], str(current_row))
            current_cell = "%s%s" % (self.extend_alphabet[col_index], str(current_row))
            current_merge = sheet[current_cell].value
            merge_cell = current_cell
            while sheet[next_cell].value is not None:
                to_compare = sheet[next_cell].value
                if current_merge == to_compare:
                    merge = "%s:%s" % (merge_cell, next_cell)
                    to_merge[merge_cell] = merge
                else:
                    current_merge = to_compare
                    merge_cell = next_cell
                col_index += 1
                current_cell = "%s%s" % (self.extend_alphabet[col_index], str(current_row))
                next_cell = "%s%s" % (self.extend_alphabet[col_index + 1], str(current_row))
            current_row += 1

        for key in to_merge.keys():
            value = to_merge.get(key)
            sheet.merge_cells(value)

    def write_reponse_details(self, sheet, table, current_row):
        current_cell = "A%s" % current_row
        current_name = table.name[11:]
        current_name = current_name.replace("$", "")
        sheet[current_cell].value = current_name
        sheet[current_cell].fill = self.shading_style
        sheet[current_cell].font = self.__font_bold
        sheet[current_cell].alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
        sheet.merge_cells(start_column=1, end_column=1, start_row = current_row, end_row=current_row+2 + len(table.responses)*4)
        bottom_cell = "A%s" % str(current_row+2)
        sheet[bottom_cell].border = self.__thick_bottom
        for response in table.responses:
            current_cell = "B%s" % current_row
            sheet[current_cell].value = response.name
            sheet[current_cell].font = self.__font_bold
            sheet[current_cell].fill = self.shading_style
            sheet[current_cell].alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
            sheet.merge_cells(start_column=2, end_column=2, start_row=current_row, end_row = current_row+2)
            border_row = current_row+3
            current_cell = "B%s" % str(border_row)
            sheet[current_cell].fill = self.shading_style
            sheet[current_cell].border = self.__thin_bottom
            sheet.row_dimensions[border_row].height = 1
            current_row += 4
        
        current_cell = "B%s" % current_row
        sheet[current_cell].value = "Total"
        sheet[current_cell].font = self.__font_bold
        sheet[current_cell].fill = self.shading_style
        sheet[current_cell].alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
        sheet.merge_cells(start_column=2, end_column=2, start_row=current_row, end_row = current_row+2)    

    def write_cross_details(self, sheet, table, current_row):
        self.banner_totals = {}
        for response in table.responses:
            current_col_index = 2
            current_cell = "%s%s" % (self.extend_alphabet[current_col_index], str(current_row))
            sheet[current_cell].value = response.population
            sheet[current_cell].alignment = self.__align
            sheet[current_cell].font = self.__font_reg
            current_cell = "%s%s" % (self.extend_alphabet[current_col_index], str(current_row + 1))
            sheet[current_cell].value = response.percentage
            sheet[current_cell].alignment = self.__align
            sheet[current_cell].font = self.__font_reg
            sheet[current_cell].number_format = '0%'
            current_cell = "%s%s" % (self.extend_alphabet[current_col_index], str(current_row + 2))
            sheet[current_cell].value = response.sig_details
            sheet[current_cell].alignment = self.__align
            sheet[current_cell].font = self.__font_reg
            current_cell = "%s%s" % (self.extend_alphabet[current_col_index], str(current_row + 3))
            sheet[current_cell].border = self.__thin_bottom
            for banner_pt in response.banner_pts:
                current_col_index += 1
                current_cell = "%s%s" % (self.extend_alphabet[current_col_index], str(current_row))
                sheet[current_cell].value = banner_pt.population
                sheet[current_cell].alignment = self.__align
                sheet[current_cell].font = self.__font_reg
                current_cell = "%s%s" % (self.extend_alphabet[current_col_index], str(current_row + 1))
                sheet[current_cell].value = banner_pt.percentage
                sheet[current_cell].alignment = self.__align
                sheet[current_cell].font = self.__font_reg
                sheet[current_cell].number_format = '0%'
                current_cell = "%s%s" % (self.extend_alphabet[current_col_index], str(current_row + 2))
                sheet[current_cell].value = banner_pt.sig_details
                if banner_pt.sig_details.isupper() is True:
                    sheet[current_cell].fill = self.highlight_style
                sheet[current_cell].alignment = self.__align
                sheet[current_cell].font = self.__font_reg
                current_cell = "%s%s" % (self.extend_alphabet[current_col_index], str(current_row + 3))
                sheet[current_cell].border = self.__thin_bottom

            current_row += 4

        self.__total_row = current_row

        current_cell = "C%s" % current_row
        sheet[current_cell].value = table.base_size
        sheet[current_cell].font = self.__font_reg
        sheet[current_cell].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    
        current_cell = "C%s" % str(current_row+1)
        sheet[current_cell].value = 1
        sheet[current_cell].number_format = '0%'
        sheet[current_cell].font = self.__font_reg
        sheet[current_cell].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        column_no = 3
        index = 0
        for population in table.total_row.populations:
            current_cell = "%s%s" % (self.extend_alphabet[column_no], current_row)
            sheet[current_cell].value = population
            sheet[current_cell].font = self.__font_reg
            sheet[current_cell].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        
            current_cell = "%s%s" % (self.extend_alphabet[column_no], str(current_row+1))
            sheet[current_cell].value = table.total_row.percentages[index]
            sheet[current_cell].number_format = '0%'
            sheet[current_cell].font = self.__font_reg
            sheet[current_cell].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            column_no += 1
            index += 1
                  
        current_row += 3
        return current_row
             
    def sig_definition(self, sheet, current_row, table):
        significant_definitions = table.sig_desc
        for definition in significant_definitions:
            current_cell = "A%s" % current_row
            sheet[current_cell].value = definition
            sheet[current_cell].font = Font(name = 'Arial', size = 6)
            current_row += 1

    def place_table_border(self, sheet, current_row, table):
        number_of_cols = table.count_banner_pts + 3
        index = 0
        while index < number_of_cols:
            current_cell = "%s%s" % (self.extend_alphabet[index], current_row)
            sheet[current_cell].border = self.__thin_top
            index += 1

        start_row = 1
        while start_row < current_row:
            current_cell = "%s%s" % (self.extend_alphabet[index], start_row)
            sheet[current_cell].border = self.__thin_left
            start_row += 1

    