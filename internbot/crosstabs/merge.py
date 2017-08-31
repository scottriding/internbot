import csv
import openpyxl

class XLSXMerger(object):

    def __init__(self):
        self.sheet_names = []

    def merge(self, path_to_folder, path_to_output, path_to_template):
        path_to_csv = path_to_folder + 'Tables to run.csv'
        output_workbook = openpyxl.load_workbook(path_to_template)
        toc_sheet = output_workbook['TOC']
        self.write_table_of_contents(path_to_csv, toc_sheet)
        self.merge_crosstabs(path_to_folder, output_workbook)
        output_workbook.save(path_to_output + 'test2.xlsx')

    def write_table_of_contents(self, path_to_tables, toc_sheet):
        table_index = 'C'
        question_name = 'D'
        question_prompt = 'E'
        base = 'F'
        iteration = 10
        with open(path_to_tables, 'rb') as csvfile:
            file = csv.DictReader(csvfile, quotechar = '"')
            for table_info in file:
                toc_sheet[table_index + str(iteration)] = table_info['TableIndex']
                sheet_name = table_info['VariableName'].translate(None,"$")
                toc_sheet[question_name + str(iteration)] = sheet_name
                toc_sheet[question_prompt + str(iteration)] = table_info['Title']
                if table_info['Base'] is not '':
                    toc_sheet[base + str(iteration)] = table_info['Base']
                iteration += 1
                self.sheet_names.append(sheet_name)

    def merge_crosstabs(self, path_to_folder, output_workbook):
        total_index = len(self.sheet_names)
        iteration = 1
        while iteration <= total_index:
            path_to_sheet = path_to_folder + str(iteration) + '.xlsx'
            temporary_workbook = openpyxl.load_workbook(path_to_sheet)
            temporary_sheet = temporary_workbook[str(iteration)]
            name = self.sheet_names[iteration-1]
            new_sheet = output_workbook.create_sheet(name)
            for row in temporary_sheet.rows:
                for cell in row:
                    new_sheet[cell.coordinate] = cell.value
                    new_sheet[cell.coordinate].font = cell.font
                    new_sheet[cell.coordinate].border = cell.border
                    new_sheet[cell.coordinate].fill = cell.fill
            iteration += 1
        