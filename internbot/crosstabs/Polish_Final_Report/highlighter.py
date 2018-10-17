from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from cell import Cells, Cell, PercentageCell, PopulationCell, SignificantMarker

class Highlighter(object):

    def __init__(self, path_to_xlsx, is_trended_amazon = False):
        self.trended = is_trended_amazon
        self.workbook = load_workbook(path_to_xlsx + '/Unhighlighted.xlsx')
        self.groups = {}
        self.responses = {}
        self.__cells = Cells()
        self.highlight_style = PatternFill("solid", fgColor="C00201")
        self.font_style = Font(name='Arial', size=10, color='ffffff')
    
    def highlight(self, path_to_output):
        for sheet in self.workbook.worksheets:
            if sheet.title != 'TOC':
                self.parse_rows(sheet)
                self.parse_columns(sheet)
                self.create_cells(sheet)
                self.assign_significant(sheet)
                self.highlight_significant(sheet)
            self.groups = {}
            self.responses = {}
            self.__cells = Cells()
        self.workbook.save(path_to_output + '/Highlighted.xlsx')
        
    def parse_rows(self, sheet):
        label_column = sheet['A']
        response_column = sheet['B']
        current_row = 3
        for label_cell in label_column:
            if label_cell.row > 3 and \
                label_cell.value is not None and \
                label_cell.value != "Results are based on two-sided tests. For each significant pair, the key of the category with the smaller column proportion appears in the category with the larger column proportion.\n Significance level for upper case letters (A, B, C): .05" and \
                label_cell.value != "a. This category is not used in comparisons because its column proportion is equal to zero or one." and \
                label_cell.value != "b. This category is not used in comparisons because the sum of case weights is less than two." and \
                label_cell.value != "Comparisons of Column Proportions":
                current_label = label_cell.value
                for cell in response_column:
                    if cell.value != "Sigma":
                        if cell.row > current_row:
                            if cell.value is not None:
                                response_label = current_label+cell.value
                                response_label.
                                if response_label not in self.responses.keys():
                                    self.responses[response_label] = cell.row
                                elif response_label in self.responses.keys():
                                    temp_storage = self.responses.get(response_label)
                                    self.responses[response_label] = [temp_storage]
                                    self.responses[response_label].append(cell.row)
                            current_row += 1
                    #response_label = current_label+cell.value
                    #if self.responses[response_label] not in self.responses.keys():
                        #self.responses[response_label] = cell.row
                    else:
                        current_row += 1
                        break
#                     elif
#                 cell.value not in self.responses.keys():
#                     self.responses[cell.value] = cell.row
#                 elif cell.value is not None and \
#                     cell.value in self.responses.keys():
#                     temp_storage = self.responses.get(cell.value)
#                     self.responses[cell.value] = [temp_storage]
#                     self.responses[cell.value].append(cell.row)

    def parse_columns(self, sheet):
        if self.trended == True:
            group_row = sheet['5']
        else:
            group_row = sheet['3']
        iteration = 1
        for cell in group_row:
            if cell.value is not None:
                if cell.value in self.groups.keys():
                    new_name = "%s_%s" % (cell.value, iteration)
                    iteration += 1
                    self.groups[new_name] = cell.column
                else:
                    self.groups[cell.value] = cell.column

    def create_cells(self, sheet):
        for group in self.groups:
            for label in self.responses:
                try:
                    location1 = str(self.groups[group]) + str(self.responses[label][0])
                    location2 = str(self.groups[group]) + str(self.responses[label][0] + 1)
                    location3 = str(self.groups[group]) + str(self.responses[label][1])
                    
                    self.__cells.add(PopulationCell(label, group, location1))
                    self.__cells.add(PercentageCell(label, group, location2))
                    self.__cells.add(SignificantMarker(label, group, location3))
                except TypeError:
                    print sheet.title
                    for key in self.responses.keys():
                        print key
                    

    def assign_significant(self, sheet):
        for cell in self.__cells:
            if cell.type == 'SignificantMarker':
                sheet_cell = sheet[cell.location].value
                if sheet_cell is not None and sheet_cell.isupper() is True:
                    match_cell = self.__cells.matching_cells(cell)
                    match_cell.is_significant = True

    def highlight_significant(self, sheet):
        for cell in self.__cells:
            if cell.is_significant is True:
                sheet[cell.location].fill = self.highlight_style
                sheet[cell.location].font = self.font_style